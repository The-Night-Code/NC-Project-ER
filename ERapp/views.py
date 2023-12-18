from django.shortcuts import render
from django.http import HttpResponse,HttpResponseRedirect
from django.forms import inlineformset_factory
from django.contrib.auth.models import User

from django.shortcuts import redirect , get_object_or_404
from django.contrib.auth import authenticate, login , logout,update_session_auth_hash
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from django.views.generic.edit import CreateView
from django.db.models import Max,Count
from django.db.models.functions import TruncDate
from .models import MyModel, UpdatedXLSXFile 
from .models import ImageModel,USER,TableData001,kizeo_model,message_box_1,kizeo_model_Pieces,AI_or_AGENT
from .models import file_table_auditV1,file_table_auditV2,file_table_auditV3,file_table_vt,file_table_auditFinal,Activities_audit,file_table_AdA,file_table_comm
from .models import Activities_be,Activities_audit
from django.template.loader import render_to_string
from django.utils import timezone

from django.core.mail import send_mail
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt,csrf_protect

from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent.parent

# Create your views here.
import os
import random
import string
from datetime import  date,timedelta
import calendar
import pytz
# Get the current time in the GMT+1 time zone
tz_gmt_plus_1 = pytz.timezone('Europe/Berlin') 

## generate xlsx
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from io import BytesIO

XLSX_TEMPLATE_FILE_PATH = BASE_DIR/ 'ERapp/static/sys_files/KiFile.xlsx'


## generate pdf
from jinja2 import Environment, FileSystemLoader
from reportlab.pdfgen import canvas
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
PDF_TEMPLATE_FILE_PATH = BASE_DIR/ 'ERapp/static/sys_files/template_pdf.html'

## backup
import tempfile
import shutil
import zipfile
import subprocess


#from ..ER_project.settings import EMAIL_HOST_USER
FROM_EMAIL="guhgi155@gmail.com"
EMAIL_SENDER = FROM_EMAIL
formT="/"
formK="/formK/"
formK111="formK"
link2="/agentimmo/"
beaudit="/beaudit/"
redirect_page_be="/be/"
VT="/VT/"

def main_Page(request):
    return render(request,'html/mainPage.html')
def Home(request):

    
    #return render(request,'html/home.html',{"name":"night","username":"nightcode"})
    return render(request,'html/home.html')
    
def send_email(subject,msg,receiver_email):
    send_mail(
    subject,
    msg,
    EMAIL_SENDER,
    [receiver_email],
    fail_silently=False,) 


@csrf_protect
def LoginU(request):
    if request.method == "POST":
        
        usernameU =request.POST.get('username')
        passwordU =request.POST.get('password')
        
        user = authenticate(username=usernameU,password=passwordU)
        if user is not None:
            login(request,user)
            return redirect("/")
    return render(request,'html/mainPage.html')



@csrf_protect
def SignupU(request):
    
    
    pl=False
    pm=False
    pc=False
    pn=False
    per=False
    if request.method == 'POST':
        username =request.POST.get('username')
        emailU =request.POST.get('email')
        pw1 =request.POST.get('password1')
        pw2 =request.POST.get('password2')
        
        per = True
        
        if str(pw1) == str(pw2):
            pm=True
            for i in str(pw1):
                # if string has letter
                if i in "abcdefghijklmnopqrstuvwxyz" or i in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
                    pc = True
                # if string has number
                if i in "0123456789":
                    pn = True

        if len(str(pw1)) >= 8 :
            pl=True
            
    
        if pl and pm and pc and pn :
            data = User.objects.create_user(username=username, email=emailU , password=pw1)
            data.save()
            return render(request, 'html/home.html')

        
    return render(request, 'html/signup.html',{"per":per,"pl":pl,"pm":pm,"pc":pc,"pn":pn})
    
    
@login_required
def LogoutU(request):
    logout(request)
    return redirect("main_page")
    #return render(request,'html/login.html')

@login_required
def ProfileU(request):
    change_password_state=False
    #user_Id="3"
    user_L=request.user
    Submit_Upload_image=request.POST.get("Submit_Upload_image")
    remove_profile_image=request.POST.get("remove_profile_image")
    change_password_button = request.POST.get("change_password")
    # Upload Profile Image
    if request.method == 'POST' :
        old_image = user_L.profile_pic
        if Submit_Upload_image=="Submit_Upload_image" and request.FILES['my_uploaded_image']:
            imagefile = request.FILES['my_uploaded_image']
            user_L.profile_pic =imagefile 
            user_L.save(update_fields=['profile_pic'])
            
            if len(old_image) > 0 and old_image and os.path.exists(old_image.path):
                try:
                    os.remove(old_image.path)    
                except:
                    pass
            return redirect("/profile/")
            
        if remove_profile_image=="remove_profile_image"  :
            
            user_L.profile_pic = 'images/sys_image/default_user_avatar.png'
            user_L.save(update_fields=['profile_pic'])
            if old_image.path != '/media/images/sys_image/default_user_avatar.png':
                if len(old_image) > 0 and os.path.exists(old_image.path) :
                    try:
                        os.remove(old_image.path)
                    except:
                        pass
            return redirect("/profile/")
            
 
            
        if change_password_button == "submit":
            newPassword = str(request.POST.get("newPassword"))
            renewPassword = str(request.POST.get("renewPassword"))
            if newPassword == renewPassword and newPassword!=" " and newPassword and len(newPassword)>=8:
                subject = 'Votre mot de passe a été changé'
                message = f'Email: {user_L.email} Mot de passe: {newPassword} '
                from_email = ''
                recipient_list = [user_L.email]
                send_mail(subject, message, from_email, recipient_list) 
                
                user_= request.user
                user_.set_password(newPassword)
                user_.save()
                update_session_auth_hash(request, user_)
                change_password_state=True
                
        

    profile_image =user_L.profile_pic
    if profile_image :
        #profile_image="{% static 'image/default_user_avatar.png' %}"
        profile_image =user_L.profile_pic.url
    

    return render(request,'html/profileU.html',{'profile_image':profile_image,'change_password_state':change_password_state})

def forgot_password(request):
    msg = False
    email = request.POST.get('email')
    change_password_state = False
    if request.method == 'POST' :
        if USER.objects.get(email=email):
            user_L=USER.objects.get(email=email)
            newPassword = generate_random_string(12)
            
            subject = 'Votre mot de passe a été changé'
            message = f'Email: {user_L.email} Mot de passe: {newPassword} '
            receiver_email = [user_L.email]
            send_email(subject,message,receiver_email)
             
            
            user_= USER.objects.get(email=email)
            user_.set_password(newPassword)
            user_.save()
            update_session_auth_hash(request, user_)
            change_password_state=True

    
    return render(request,'html/forgot_password.html',{'change_password_state': change_password_state})

def showimage(request):
    profile_img = ImageModel.objects.get(user_id="n123") 
    return render(request,'html/showimage.html',{'profile_image': profile_img})






def generate_random_string(length):
    characters = string.ascii_letters + string.digits  # You can customize this for your needs
    random_string = ''.join(random.choice(characters) for _ in range(length))
    random_string=random_string.replace("%20","")
    random_string=random_string.replace(" ","")
    return random_string


@login_required
def add_f_to_table_view_2(request,myID1,column_name_type,button_edit_data_on_table,Activity_table):
    l1=f"table_{column_name_type}[]"
    file_table = ModelByColumn(column_name_type)
    files_date_for_response=[]
    I_icon_class=""
    files_date_for_response+=[{'file_id':"button_edit_data_on_table",
                                                        'file_save':"file",
                                                        'file_format':"format_file",
                                                        'file_name':"file.name",
                                                        'file_index':"1111",
                                                        'column':"column_name_type",
                                                        'I_icon_class':"I_icon_class",},]
    files_date_for_response+=[{'file_id':"button_edit_data_on_table",
                                                        'file_save':"file2",
                                                        'file_format':"format_file2",
                                                        'file_name':"file.name2",
                                                        'file_index':"1111",
                                                        'column':"column_name_type",
                                                        'I_icon_class':"I_icon_class",},]
    
    for file in request.FILES.getlist(l1, []):
        format_file=file.name.split(".")[1]
        if format_file in ['jpg','png','jpeg','heic']:
            format_file="image"
            I_icon_class="bi bi-file-earmark-image"
        elif format_file in ['doc','docx']:
            format_file="word"
            I_icon_class="bi bi-file-earmark-word"
        elif format_file in ['xls','xlsm']:
            format_file="excel" 
            I_icon_class="ri-file-excel-2-line"
        elif format_file in ['pdf']:
            format_file="pdf" 
            I_icon_class="bi bi-file-earmark-pdf"
        elif format_file in ['pz2']:
            format_file="pz2" 
            I_icon_class="bi bi-file-ppt"
        else:
            I_icon_class="ri-file-line"
        
        
        if not file_table.objects.filter(file_id=button_edit_data_on_table, file_format=format_file,file_name=file.name):
            Activities_audit.objects.create(
                    Activity_id=generate_random_string(10),
                    Activity_user = f"{request.user.last_name} {request.user.first_name}",
                    Activity_user_email = request.user.email,
                    Activity_table=Activity_table,
                    Activity_project_id = button_edit_data_on_table,
                    Activity_before =f"le fichier {file.name}" ,
                    Activity_after = column_name_type ,
                    Activity_add=True
                )
            file_table.objects.create(
                    file_id = button_edit_data_on_table,
                    file_name = file.name,
                    file_save = file,
                    file_format =format_file
                )
            
            obj_by_id = get_object_or_404(file_table,file_id=button_edit_data_on_table, file_format=format_file,file_name=file.name,file_save = file)
            #obj_by_id = file_table.objects.filter(file_id=button_edit_data_on_table, file_format=format_file,file_name=file.name,file_save = file)
            #GET_file_index = getattr(obj_by_id,'file_index')
            column_name_type_new = column_name_type
            if column_name_type == "VT":
                column_name_type_new="vt"
            files_date_for_response.append({'file_id':button_edit_data_on_table,
                                        'file_save':file,
                                        'file_format':format_file,
                                        'file_name':file.name,
                                        'file_index':"1111",
                                        'column':column_name_type,
                                        'I_icon_class':I_icon_class,})
            
    return files_date_for_response
    
            
@login_required
def add_f_to_table_view(request,myID1,column_name_type,button_edit_data_on_table,Activity_table):
    l1=f"table_{column_name_type}_{button_edit_data_on_table}"
    file_table = ModelByColumn(column_name_type)
    for file in request.FILES.getlist(l1):
        format_file=file.name.split(".")[1]
        if format_file in ['jpg','png','jpeg','heic']:
            format_file="image"
        if format_file in ['doc','docx']:
            format_file="word"
        if format_file in ['xls','xlsm']:
            format_file="excel" 
            
        
        
        if not file_table.objects.filter(file_id=button_edit_data_on_table, file_format=format_file,file_name=file.name):
            Activities_audit.objects.create(
                    Activity_id=generate_random_string(10),
                    Activity_user = f"{request.user.last_name} {request.user.first_name}",
                    Activity_user_email = request.user.email,
                    Activity_table=Activity_table,
                    Activity_project_id = button_edit_data_on_table,
                    Activity_before =f"le fichier {file.name}" ,
                    Activity_after = column_name_type ,
                    Activity_add=True
                )
            file_table.objects.create(
                    file_id = button_edit_data_on_table,
                    file_name = file.name,
                    file_save = file,
                    file_format =format_file
                )
@login_required
def table_view_2(request):
    #Activity_id,Activity_user,Activity_table,Activity_in,Activity_before,Activity_after
    user_=request.user
    
    if request.method == 'POST':
        #return redirect("/Accueil/")
        redirect_page = request.POST.get("redirect_page")
        myID1=request.POST.get("myid1")
        column1=request.POST.get("col_type1")
        Activity_table=''
        file_uploaded_state=False
        files_date_for_response=[]
        button_edit_data_on_table=request.POST.get("cellId_new")
        file_Project_id =request.POST.get("cellId_new")
        if button_edit_data_on_table:

            try:
                obj_by_id = get_object_or_404(TableData001,cell_id=str(button_edit_data_on_table))  # TableData001.objects.get(cell_id=str(button_edit_data_on_table))

                table_index=[   {'column_name':'firstname','name2':'Prénom'},
                                {'column_name':'lastname', 'name2':'Nom'},
                                {'column_name':'address',  'name2':'Adresse'},
                                {'column_name':'email',    'name2':'Email'},
                                {'column_name':'num',      'name2':'N° de tél'},
                                
                                {'column_name':'tp',       'name2':'Travaux à préconiser'},
                                {'column_name':'cofrac',   'name2':'Cofrac'},
                                {'column_name':'auditeur', 'name2':'Auditeur'},
                                {'column_name':'paiement', 'name2':'Paiement'},
                                #{'column_name':'precaite', 'name2':'Précarite'}
                                
                                {'column_name':'etat',     'name2':'État'},
                                ]
                
                
                if obj_by_id.be:
                    Activity_table="Bureau d'étude"
                elif obj_by_id.ai:
                    Activity_table="Agent immobilier"
                
                for l in table_index :
                    name = l['column_name']
                    name2 = l['name2']
                    re_page = False
                    
                    
                    if name == 'paiement':
                        #Activity_before = str()
                        Activity_after = request.POST.get(f"table_paiement")
                        Activity_before_=""
                        Activity_after_=""
                        Activity_after_="impayé"
                        AF_af = False                    
                        if Activity_after == "True":
                            Activity_after_="payé"
                            AF_af = True
                        else: 
                            if not Activity_after == "True":
                                Activity_after_="impayé"
                                AF_af = False
                            
                        if getattr(obj_by_id, name) is True:
                            Activity_before_="payé"
                        else  :
                            Activity_before_="impayé"
                
                        if Activity_before_ != Activity_after_:
                            Activities_audit.objects.create(Activity_id=generate_random_string(10),
                                                            Activity_user=f"{user_.last_name} {user_.first_name}",
                                                            Activity_user_email=user_.email,
                                                            Activity_table=Activity_table,
                                                            Activity_project_id=str(button_edit_data_on_table),
                                                            Activity_name=name2,
                                                            Activity_before=Activity_before_,
                                                            Activity_after=Activity_after_,
                                                            Activity_edit=True)
                        setattr(obj_by_id, 'paiement', AF_af)
                        obj_by_id.save()
                    else:
                        
                        Activity_before = str(getattr(obj_by_id, name) )+ ""
                        Activity_after = str(request.POST.get(f"table_{name}"))+ ""
                        if Activity_before != Activity_after:
                            if name == 'etat':
                                re_page=True
                                if (Activity_after == "Fini" or Activity_after == "fini") and getattr(obj_by_id, 'Modification_Faite_time_checker_2') is True:
                                    setattr(obj_by_id, 'Modification_Faite_by_user', user_.email)
                                    setattr(obj_by_id, 'Modification_Faite_time_checker_2', False)
                                    setattr(obj_by_id, 'Modification_Faite_time', timezone.now())
                                    
                            Activities_audit.objects.create(Activity_id=generate_random_string(10),
                                                            Activity_user=f"{user_.last_name} {user_.first_name}",
                                                            Activity_user_email=user_.email,
                                                            Activity_table=Activity_table,
                                                            Activity_project_id=str(button_edit_data_on_table),
                                                            Activity_name=name2,
                                                            Activity_before=Activity_before,
                                                            Activity_after=Activity_after,
                                                            Activity_edit=True)
                            
                            
                        setattr(obj_by_id, name, Activity_after)
                        obj_by_id.save()
                
                   
                        
                    
                    

                
                if request.POST.get(f"table_etat") == "Envoye" and not obj_by_id.Envoye_time_checker :
                    obj_by_id.Envoye_time_checker = True
                    obj_by_id.Envoye_time = timezone.now()
                    obj_by_id.Envoye_by_user = user_.email
                    obj_by_id.save(update_fields=['Envoye_time_checker','Envoye_time','Envoye_by_user'])
                
                if request.POST.get(f"table_etat") =="Fini" and not obj_by_id.fini_time_checker :
                    obj_by_id.fini_time_checker = True
                    obj_by_id.fini_time = timezone.now()
                    obj_by_id.fini_by_user = user_.email
                    obj_by_id.save(update_fields=['fini_time_checker','fini_time','fini_by_user'])
                    
                    
            except TableData001.DoesNotExist:
                pass
            
            VT_files_added = 0 
            auditV1_files_added = 0
            auditV2_files_added = 0
            auditV3_files_added = 0
            auditFinal_files_added = 0
            files_date_for_response=[]
            column_name_types = ["VT", "auditV1", "auditV2", "auditV3", "auditFinal"]
            for column_name_type in column_name_types:
                files_uploaded = request.FILES.getlist(f"table_{column_name_type}[]")
                if files_uploaded:
                    file_uploaded_state = True
                    l1=f"table_{column_name_type}[]"
                    file_table = ModelByColumn(column_name_type)
                    I_icon_class=""

                    
                    for file in request.FILES.getlist(l1, []):
                                                
                        format_file=file.name.split(".")[1]
                        if format_file not in ['jpg','png','jpeg','heic','doc','docx','xls','xlsm','pdf','pz2']:
                            I_icon_class="ri-file-line"
                            
                        if format_file in ['jpg','png','jpeg','heic']:
                            format_file="image"
                            I_icon_class="bi bi-file-earmark-image"
                        if format_file in ['doc','docx']:
                            format_file="word"
                            I_icon_class="bi bi-file-earmark-word"
                        if format_file in ['xls','xlsm']:
                            format_file="excel" 
                            I_icon_class="ri-file-excel-2-line"
                        if format_file in ['pdf']:
                            format_file="pdf" 
                            I_icon_class="bi bi-file-earmark-pdf"
                        if format_file in ['pz2']:
                            format_file="pz2" 
                            I_icon_class="bi bi-file-ppt"
                        
                        
                        
                        if not file_table.objects.filter(file_id=button_edit_data_on_table, file_name=file.name,file_removed=False):
                            column_name_type_new = column_name_type
                            if column_name_type == "VT":
                                VT_files_added +=1
                                column_name_type_new="vt"
                                
                            if column_name_type == "auditV1":
                                auditV1_files_added +=1
                            
                            if column_name_type == "auditV2":
                                auditV2_files_added +=1
                            
                            if column_name_type == "auditV3":
                                auditV3_files_added +=1
                                
                            if column_name_type == "auditFinal":
                                auditFinal_files_added +=1
                            try:
                                Activities_audit.objects.create(
                                        Activity_id=generate_random_string(10),
                                        Activity_user = f"{request.user.last_name} {request.user.first_name}",
                                        Activity_user_email = request.user.email,
                                        Activity_table=Activity_table,
                                        Activity_project_id = button_edit_data_on_table,
                                        Activity_before =f"le fichier {file.name}" ,
                                        Activity_after = column_name_type ,
                                        Activity_add=True
                                    )

                                file_table.objects.create(
                                        file_id = button_edit_data_on_table,
                                        file_name = file.name,
                                        file_save = file,
                                        file_format =format_file
                                    )
                    
                                files_date_for_response.append({'file_id':button_edit_data_on_table,
                                                            'file_save_url':"file.name",
                                                            'file_format':format_file,
                                                            'file_name':file.name,
                                                            'file_index':"",
                                                            'column':column_name_type_new,
                                                            'I_icon_class':I_icon_class,})
                                
                            except:
                                pass
            
            
                        
            VT_col_n = "VT"
            auditV1_col_n = "auditV1"
            auditV2_col_n = "auditV2"
            auditV3_col_n = "auditV3"
            auditFinal_col_n = "auditFinal"
            files_added_list3=[{'VT_files_added':VT_files_added,'file_id':file_Project_id,'col_n':"VT"},
                              {'auditV1_files_added':auditV1_files_added,'file_id':file_Project_id,'col_n':"auditV1"},
                              {'auditV2_files_added':auditV2_files_added,'file_id':file_Project_id,'col_n':"auditV2"},
                              {'auditV3_files_added':auditV3_files_added,'file_id':file_Project_id,'col_n':"auditV3"},
                              {'auditFinal_files_added':auditFinal_files_added,'file_id':file_Project_id,'col_n':"auditFinal"}]
            files_added_list=[{'files_added':VT_files_added,'file_id':file_Project_id,'col_n':"VT"},
                              {'files_added':auditV1_files_added,'file_id':file_Project_id,'col_n':"auditV1"},
                              {'files_added':auditV2_files_added,'file_id':file_Project_id,'col_n':"auditV2"},
                              {'files_added':auditV3_files_added,'file_id':file_Project_id,'col_n':"auditV3"},
                              {'files_added':auditFinal_files_added,'file_id':file_Project_id,'col_n':"auditFinal"}]
            response_date={
                're_page':re_page,
                'cellId_new':button_edit_data_on_table,
                'file_uploaded_state':file_uploaded_state,
                'files_date_for_response':files_date_for_response,
                'files_added_list':files_added_list,
            }
            
            return JsonResponse(response_date)
    return JsonResponse({'status': 'error'})





@login_required
def Auditeur_Accueil(request):
    user_=request.user
    data = TableData001.objects.all()
    activities_audit = Activities_audit.objects.order_by("-Activity_date")[:6]
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    first_day_of_year = date(today.year, 1, 1)
    first_day_of_month = date(today.year, today.month, 1)

    client_count_added_today = TableData001.objects.filter(creation_time__range=(yesterday,today)).count()
    client_count_added_month = TableData001.objects.filter(creation_time__range=(first_day_of_month, timezone.now())).count()
    client_count_added_year= TableData001.objects.filter(creation_time__range=(first_day_of_year, timezone.now())).count()

    client_count_envoye_today = TableData001.objects.filter(Envoye_time__range=(yesterday,today)).count()
    client_count_envoye_month = TableData001.objects.filter(Envoye_time__range=(first_day_of_month, timezone.now())).count()
    client_count_envoye_year= TableData001.objects.filter(Envoye_time__range=(first_day_of_year, timezone.now())).count()
    
    client_count_fini_today = TableData001.objects.filter(fini_time__range=(yesterday,today)).count()
    client_count_fini_month = TableData001.objects.filter(fini_time__range=(first_day_of_month,timezone.now())).count()
    client_count_fini_year= TableData001.objects.filter(fini_time__range=(first_day_of_year, timezone.now())).count()
    
    client_count_Modification_Faite_today = TableData001.objects.filter(Modification_Faite_time__range=(yesterday,today)).count()
    client_count_Modification_Faite_month = TableData001.objects.filter(Modification_Faite_time__range=(first_day_of_month,timezone.now())).count()
    client_count_Modification_Faite_year= TableData001.objects.filter(Modification_Faite_time__range=(first_day_of_year, timezone.now())).count()
    
    fini_result={}
    envoye_result={}
    Modification_Faite_result={}
    days_of_week = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi','dimanche']
    
    # Get the current date
    current_date = timezone.now().date()
    # Get the current day of the week (0 for Monday, 1 for Tuesday, etc.)
    current_weekday = current_date.weekday()

    # Loop through the past N days, where N is the current day of the week
    k=int(current_weekday)
    for day_offset in range(current_weekday + 1):
        # Calculate the date range for each day
        start_date = current_date - timedelta(days=day_offset)
        end_date = start_date + timedelta(days=1) - timedelta(seconds=1)

        # Filter and count records for the specified date range
        count1 = TableData001.objects.filter(fini_time__range=(start_date, end_date),fini_time_checker=True,fini_by_user=user_.email).count()
        count2 = TableData001.objects.filter(Envoye_time__range=(start_date, end_date),Envoye_time_checker=True,Envoye_by_user=user_.email).count()
        count3 = TableData001.objects.filter(Modification_Faite_time__range=(start_date, end_date),Modification_Faite_time_checker=True,Modification_Faite_by_user=user_.email).count()
        
        # Store the count in the dictionary
        fini_result[days_of_week[k]] = count1
        envoye_result[days_of_week[k]] = count2
        Modification_Faite_result[days_of_week[k]] = count3
        k-=1
        
    #final_result={'lundi':re1 , 'mardi':re2}
    return render(request, "html/Auditeur_main_page.html",{'data':data,
                                                        'activities_audit':activities_audit,
                                                        'client_count_added_today':client_count_added_today,
                                                        'client_count_added_month':client_count_added_month,
                                                        'client_count_added_year':client_count_added_year,
                                                        
                                                        'client_count_envoye_today':client_count_envoye_today,
                                                        'client_count_envoye_month':client_count_envoye_month,
                                                        'client_count_envoye_year':client_count_envoye_year,
                                                        
                                                        'client_count_fini_today':client_count_fini_today,
                                                        'client_count_fini_month':client_count_fini_month,
                                                        'client_count_fini_year':client_count_fini_year,
                                                        
                                                        'fini_result':fini_result,
                                                        'envoye_result':envoye_result,
                                                        'Modification_Faite_result':Modification_Faite_result,
                                                        
                                                        'client_count_Modification_Faite_today':client_count_Modification_Faite_today,
                                                        'client_count_Modification_Faite_month':client_count_Modification_Faite_month,
                                                        'client_count_Modification_Faite_year':client_count_Modification_Faite_year,
                                                        })

@login_required
def audit_pages(request,redirect_page,html_page):
    
    

    data = TableData001.objects.all().order_by('-creation_time')
    
    col_count = data.count()
    # Get unique column names from the TableData model
    column_names = TableData001._meta.get_fields()
    
    datafiles_VT = file_table_vt.objects.filter(file_removed=False)
    countfiles_VT = file_table_vt.objects.filter(file_removed=False).count()
    datafiles_AuditV1 = file_table_auditV1.objects.filter(file_removed=False)
    countfiles_AuditV1 = file_table_auditV1.objects.filter(file_removed=False).count()
    datafiles_AuditV2 = file_table_auditV2.objects.filter(file_removed=False)
    countfiles_AuditV2 = file_table_auditV2.objects.filter(file_removed=False).count()
    datafiles_AuditV3 = file_table_auditV3.objects.filter(file_removed=False)
    countfiles_AuditV3 = file_table_auditV3.objects.filter(file_removed=False).count()
    datafiles_AuditFinal = file_table_auditFinal.objects.filter(file_removed=False)
    countfiles_AuditFinal = file_table_auditFinal.objects.filter(file_removed=False).count()
    message_box_01 = message_box_1.objects.all()
    #user_profile_image = USER.objects.all()#values_list('first_name','last_name','email','profile_pic')

    table_index=[{'index':1,'state':"A realiser"},
                 {'index':2,'state':"En cours"},
                 {'index':3,'state':"A modifier"},
                 {'index':4,'state':"Modification Faite"},
                 {'index':5,'state':"Reclamation"},
                 {'index':6,'state':"Reclamation Faite"},
                 {'index':7,'state':"Envoye"},
                 {'index':8,'state':"Annule"},
                 {'index':9,'state':"Fini"}]
    auditeur=[]
    for audi in USER.objects.all():
        audi_role = str(audi.role)
        if "auditeur" in audi_role : #auditeur
            auditeur+=[{'profile_pic':audi.profile_pic,'email':audi.email,'first_name':audi.first_name,'last_name':audi.last_name}]
            #bbb=1
            
    a=[]
    for email_audi in TableData001.objects.all():
        a+=[{'email':email_audi.auditeur}]
    user_=request.user
    msg_box_tagged=[user_.email,user_.first_name,user_.last_name]
    Audi=user_.email
    list={ 'data': data ,
                                                  'col_count':col_count ,
                                                  'column_names': column_names,
                                                  'datafiles_VT': datafiles_VT ,
                                                  'countfiles_VT': countfiles_VT ,
                                                  'datafiles_AuditV1': datafiles_AuditV1 ,
                                                  'countfiles_AuditV1': countfiles_AuditV1,
                                                  'datafiles_AuditV2': datafiles_AuditV2 ,
                                                  'countfiles_AuditV2': countfiles_AuditV2,
                                                  'datafiles_AuditV3': datafiles_AuditV3 ,
                                                  'countfiles_AuditV3': countfiles_AuditV3,
                                                  'datafiles_AuditFinal':datafiles_AuditFinal,
                                                  'countfiles_AuditFinal': countfiles_AuditFinal,
                                                  'message_box_1':message_box_01,
                                                  'table_index':table_index,
                                                  'auditeur':auditeur,
                                                  'Audi':Audi,
                                                  #'a':a,
                                                  'redirect_next_page':redirect_page,
                                                  'msg_box_tagged':msg_box_tagged,
                                                  #'user_profile_image':user_profile_image
                                                  }
    return list

@login_required 
def AI_audit_ALL(request): # add row (,firstname,lastname,addressm,num,etat,tp,cofrac,paiment,agent,)
    redirect_page="/AI_audit_ALL/"
    html_page="AI_audit_ALL.html"
    

    list=audit_pages(request,redirect_page,html_page)
    return render(request, f"html/{html_page}", list)

@login_required 
def AI_audit_BY_A(request): # add row (,firstname,lastname,addressm,num,etat,tp,cofrac,paiment,agent,)
    redirect_page="/AI_audit_BY_A/"
    html_page="AI_audit_BY_A.html"
    
    
    list=audit_pages(request,redirect_page,html_page)
    return render(request, f"html/{html_page}", list)


@login_required 
def BE_audit_ALL(request): # add row (,firstname,lastname,addressm,num,etat,tp,cofrac,paiment,agent,)
    redirect_page="/BE_audit_ALL/"
    html_page="BE_audit_ALL.html"
    
    
    list=audit_pages(request,redirect_page,html_page)
    return render(request, f"html/{html_page}", list)


@login_required
def BE_audit_BY_A(request):
    redirect_page="/BE_audit_BY_A/"
    html_page="BE_audit_BY_A.html"
    
    
    
    list=audit_pages(request,redirect_page,html_page)
    return render(request, f"html/{html_page}", list)
    

@login_required
def Auditor_Task_Summary(request):   #Tableau de Bord  # Résumé des Tâches par Auditeur
    
    table_index=[{'index':1,'state':"Projet fini"},
                 {'index':2,'state':"Projet Envoyé"},
                 {'index':3,'state':"Modification Faite"}]
    
    current_date = timezone.now().date()

    # Find the first day of the current month
    first_day_current_month = current_date.replace(day=1)

    # Find the first day of the last month
    first_day_last_month = (first_day_current_month - timedelta(days=1)).replace(day=1)

    # Find the last day of the last month
    last_day_last_month = first_day_current_month - timedelta(days=1)
    
    
    
    today = timezone.now().date() + timedelta(days=1) 
    yesterday = today - timedelta(days=1)
    before_yesterday = today - timedelta(days=2)
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    
    end_of_last_week = today - timedelta(days=today.weekday() + 1)
    start_of_last_week = end_of_last_week - timedelta(days=6)
    
    start_of_month = first_day_current_month # today.replace(day=1)
    end_of_month = today #start_of_month.replace(day=1, month=start_of_month.month % 12 + 1) - timedelta(days=1)
    
    start_of_last_month = first_day_last_month # today.replace(day=1) - timedelta(days=1) 
    end_of_last_month = last_day_last_month  # start_of_last_month.replace(day=1, month=start_of_last_month.month % 12 + 1) - timedelta(days=1)

    start_of_year = today.replace(month=1, day=1)
    end_of_year = start_of_year.replace(month=12, day=31)

    data=[]
    for audi in USER.objects.all():
        audi_role = str(audi.role)
        if "auditeur" in audi_role : #auditeur
            #auditeur+=[{'email':audi.email,'first_name':audi.first_name,'last_name':audi.last_name}]
            
            s_today_fini_time = TableData001.objects.filter(fini_time__range=[yesterday, today],fini_time_checker=True,fini_by_user=audi.email).count()
            s_yesterday_fini_time = TableData001.objects.filter(fini_time__range=[before_yesterday, yesterday],fini_time_checker=True,fini_by_user=audi.email).count()
            s_this_week_fini_time = TableData001.objects.filter(fini_time__range=[start_of_week, end_of_week],fini_time_checker=True,fini_by_user=audi.email).count()
            s_last_week_fini_time = TableData001.objects.filter(fini_time__range=[start_of_last_week, end_of_last_week],fini_time_checker=True,fini_by_user=audi.email).count()
            s_this_month_fini_time = TableData001.objects.filter(fini_time__range=[start_of_month, end_of_month],fini_time_checker=True,fini_by_user=audi.email).count()
            s_last_month_fini_time = TableData001.objects.filter(fini_time__range=[start_of_last_month, end_of_last_month],fini_time_checker=True,fini_by_user=audi.email).count()
            s_this_year_fini_time = TableData001.objects.filter(fini_time__range=[start_of_year, end_of_year],fini_time_checker=True,fini_by_user=audi.email).count()
            
            s_today_Envoye_time = TableData001.objects.filter(Envoye_time__range=[yesterday, today],Envoye_time_checker=True,Envoye_by_user=audi.email).count()
            s_yesterday_Envoye_time = TableData001.objects.filter(Envoye_time__range=[before_yesterday, yesterday],Envoye_time_checker=True,Envoye_by_user=audi.email).count()
            s_this_week_Envoye_time = TableData001.objects.filter(Envoye_time__range=[start_of_week, end_of_week],Envoye_time_checker=True,Envoye_by_user=audi.email).count()
            s_last_week_Envoye_time = TableData001.objects.filter(Envoye_time__range=[start_of_last_week, end_of_last_week],Envoye_time_checker=True,Envoye_by_user=audi.email).count()
            s_this_month_Envoye_time = TableData001.objects.filter(Envoye_time__range=[start_of_month, end_of_month],Envoye_time_checker=True,Envoye_by_user=audi.email).count()
            s_last_month_Envoye_time = TableData001.objects.filter(Envoye_time__range=[start_of_last_month, end_of_last_month],Envoye_time_checker=True,Envoye_by_user=audi.email).count()
            s_this_year_Envoye_time = TableData001.objects.filter(Envoye_time__range=[start_of_year, end_of_year],Envoye_time_checker=True,Envoye_by_user=audi.email).count()
            
            s_today_MF_time = TableData001.objects.filter(Modification_Faite_time__range=[yesterday, today],Modification_Faite_time_checker=True,Modification_Faite_by_user=audi.email).count()
            s_yesterday_MF_time = TableData001.objects.filter(Modification_Faite_time__range=[before_yesterday, yesterday],Modification_Faite_time_checker=True,Modification_Faite_by_user=audi.email).count()
            s_this_week_MF_time = TableData001.objects.filter(Modification_Faite_time__range=[start_of_week, end_of_week],Modification_Faite_time_checker=True,Modification_Faite_by_user=audi.email).count()
            s_last_week_MF_time = TableData001.objects.filter(Modification_Faite_time__range=[start_of_last_week, end_of_last_week],Modification_Faite_time_checker=True,Modification_Faite_by_user=audi.email).count()
            s_this_month_MF_time = TableData001.objects.filter(Modification_Faite_time__range=[start_of_month, end_of_month],Modification_Faite_time_checker=True,Modification_Faite_by_user=audi.email).count()
            s_last_month_MF_time = TableData001.objects.filter(Modification_Faite_time__range=[start_of_last_month, end_of_last_month],Modification_Faite_time_checker=True,Modification_Faite_by_user=audi.email).count()
            s_this_year_MF_time = TableData001.objects.filter(Modification_Faite_time__range=[start_of_year, end_of_year],Modification_Faite_time_checker=True,Modification_Faite_by_user=audi.email).count()
            
            data.append({'email':audi.email,'first_name':audi.first_name,'last_name':audi.last_name,
                        
                        's_today_fini_time':s_today_fini_time,
                        's_yesterday_fini_time':s_yesterday_fini_time,
                        's_this_week_fini_time':s_this_week_fini_time,
                        's_last_week_fini_time':s_last_week_fini_time,
                        's_this_month_fini_time':s_this_month_fini_time,
                        's_last_month_fini_time':s_last_month_fini_time,
                        's_this_year_fini_time':s_this_year_fini_time,
                        
                        's_today_Envoye_time':s_today_Envoye_time,
                        's_yesterday_Envoye_time':s_yesterday_Envoye_time,
                        's_this_week_Envoye_time':s_this_week_Envoye_time,
                        's_last_week_Envoye_time':s_last_week_Envoye_time,
                        's_this_month_Envoye_time':s_this_month_Envoye_time,
                        's_last_month_Envoye_time':s_last_month_Envoye_time,
                        's_this_year_Envoye_time':s_this_year_Envoye_time,
                        
                        's_today_MF_time':s_today_MF_time,
                        's_yesterday_MF_time':s_yesterday_MF_time,
                        's_this_week_MF_time':s_this_week_MF_time,
                        's_last_week_MF_time':s_last_week_MF_time,
                        's_this_month_MF_time':s_this_month_MF_time,
                        's_last_month_MF_time':s_last_month_MF_time,
                        's_this_year_MF_time':s_this_year_MF_time,
                        })


            
    
    
    return render(request, 'html/Auditeur_state.html',{'data':data,
                                                       'table_index':table_index,} )

@login_required
def Auditor_Task_Summary_BY_A(request,auditeur_email):

    
    user_A = USER.objects.get(email=auditeur_email)
    user_Au={'firstname':user_A.first_name,
            'lastname':user_A.last_name,
            'email':user_A.email,
            'profile_pic':user_A.profile_pic}
    
    table_index_=[{'index':1,'state':"Aujourd'hui"},
                 {'index':2,'state':"Hier"},
                 {'index':3,'state':""},
                 {'index':4,'state':""},
                 {'index':5,'state':""},
                 {'index':6,'state':""},
                 {'index':7,'state':""}]
    table_index=[]
    fini_result={}
    envoye_result={}
    Modification_Faite_result={}
    days_of_week = ['day-7', 'day-6', 'day-5', 'day-4', 'day-3', 'day-2','day-1']
    
    current_date = timezone.now().date()
    # Get the current day of the week (0 for Monday, 1 for Tuesday, etc.)
    current_weekday = 6

    # Loop through the past N days, where N is the current day of the week
    k=int(current_weekday)
    for day_offset in range(current_weekday +1):
        
        # Calculate the date range for each day
        start_date = current_date - timedelta(days=day_offset)
        end_date = start_date - timedelta(days=1) 
        table_index.append({'index':k,'state':start_date})
        # Filter and count records for the specified date range
        count1 = TableData001.objects.filter(fini_time__range=(end_date, start_date),fini_time_checker=True,fini_by_user=auditeur_email)
        count2 = TableData001.objects.filter(Envoye_time__range=(start_date, end_date),Envoye_time_checker=True,Envoye_by_user=auditeur_email)
        count3 = TableData001.objects.filter(Modification_Faite_time__range=(end_date, start_date),Modification_Faite_time_checker=True,Modification_Faite_by_user=auditeur_email)
        
        # Store the count in the dictionary
        fini_result[days_of_week[k]] = count1
        envoye_result[days_of_week[k]] = count2
        Modification_Faite_result[days_of_week[k]] = count3
        k-=1
    
    
    
    
    
    
    
    
    
    
    

    day_0 = timezone.now().date() + timedelta(days=1) 
    day_1 = day_0 - timedelta(days=1)
    day_2 = day_0 - timedelta(days=2)
    day_3 = day_0 - timedelta(days=3)
    day_4 = day_0 - timedelta(days=4)
    day_5 = day_0 - timedelta(days=5)
    day_6 = day_0 - timedelta(days=6)
    day_7 = day_0 - timedelta(days=7)
    


    data=[]
    table_index_2=[]
           
    day_1_fini_time = TableData001.objects.filter(fini_time__range=[day_1, day_0],fini_time_checker=True,fini_by_user=auditeur_email)
    day_2_fini_time = TableData001.objects.filter(fini_time__range=[day_2, day_1],fini_time_checker=True,fini_by_user=auditeur_email)
    day_3_fini_time = TableData001.objects.filter(fini_time__range=[day_3, day_2],fini_time_checker=True,fini_by_user=auditeur_email)
    day_4_fini_time = TableData001.objects.filter(fini_time__range=[day_4, day_3],fini_time_checker=True,fini_by_user=auditeur_email)
    day_5_fini_time = TableData001.objects.filter(fini_time__range=[day_5, day_4],fini_time_checker=True,fini_by_user=auditeur_email)
    day_6_fini_time = TableData001.objects.filter(fini_time__range=[day_6, day_5],fini_time_checker=True,fini_by_user=auditeur_email)
    day_7_fini_time = TableData001.objects.filter(fini_time__range=[day_7, day_6],fini_time_checker=True,fini_by_user=auditeur_email)
    
    days_of_week = ['day_7', 'day_6', 'day_5', 'day_4', 'day_3', 'day_2','day_1']
    Envoye_time={}
    day_0 = timezone.now().date() + timedelta(days=1) 
    j=1
    for i in range(7):
        start_date = day_0 - timedelta(days=i)
        end_date = day_0 - timedelta(days=i+1)
        day_fini_time = TableData001.objects.filter(fini_time__range=[end_date, start_date],fini_time_checker=True,fini_by_user=auditeur_email)
        day_Envoye_time = TableData001.objects.filter(Envoye_time__range=[end_date , start_date ],Envoye_time_checker=True,Envoye_by_user=auditeur_email)
        day_Modification_Faite_result_time = TableData001.objects.filter(Modification_Faite_time__range=[end_date , start_date ],Modification_Faite_time_checker=True,Modification_Faite_by_user=auditeur_email)
        table_index_2.append({'state':end_date,
                              'index':j,
                            'fini_result':day_fini_time,
                            'envoye_result':day_Envoye_time,
                            'Modification_Faite_result':day_Modification_Faite_result_time})
        j+=1
    #day_1_ 

    data.append({'email':auditeur_email,#'first_name':audi.first_name,'last_name':audi.last_name,

                'fini_result':fini_result,
                'envoye_result':envoye_result,
                'Modification_Faite_result':Modification_Faite_result})
    
    return render(request, 'html/Auditeur_state_by_a.html',{'data':data,
                                                        'user_Au':user_Au,
                                                        'table_index':table_index,
                                                        'table_index_2':table_index_2,
                                                        'day_1_fini_time':day_1_fini_time,
                                                        'day_2_fini_time':day_2_fini_time,
                                                        'day_3_fini_time':day_3_fini_time,
                                                        'day_4_fini_time':day_4_fini_time,
                                                        'day_5_fini_time':day_5_fini_time,
                                                        'day_6_fini_time':day_6_fini_time,
                                                        'day_7_fini_time':day_7_fini_time,
                                                        } )


@login_required
def send_message(request):
    if request.method == 'POST':
        user_ = request.user
        message = request.POST.get('message')
        redirectNextPage = request.POST.get('redirectNextPage')
        #box = request.POST.get('box')
        col = request.POST.get('col')
        cellId = request.POST.get('cellId')
        msg_id=generate_random_string(10)
        if message:
            
            new_message = message_box_1.objects.create(
                                                       
                                                        message_id = msg_id,
                                                        row_id = cellId,
                                                        username =user_.first_name + "  " +  user_.last_name,
                                                        email =  user_.email,
                                                        message =message,
                                                        box = col,
                                                        profile_pic = user_.profile_pic)
            response_data = {
                'username': new_message.username,
                'message': new_message.message,
                #'messageDate': new_message.timestamp,
                'userProfilePic': request.user.profile_pic.url if request.user.profile_pic else ''  # Adjust this to your User model
            }
            return JsonResponse(response_data)
    return JsonResponse({'status': 'error'})

    

def ModelByColumn(model_by_column):
    if model_by_column == "VT" or model_by_column == "vt":
        return file_table_vt
    if model_by_column == "auditV1":
        return file_table_auditV1
    if model_by_column == "auditV2":
        return file_table_auditV2
    if model_by_column == "auditV3":
        return file_table_auditV3
    if model_by_column == "auditFinal":
        return file_table_auditFinal
    
    if model_by_column == "AdA":
        return file_table_AdA
    
    if model_by_column == "comm":
        return file_table_comm
    
@login_required   
def remove_file_from_MODELS(request):
    if request.method == 'POST':
        user_ = request.user
        user_role= user_.role
        #if  "auditeur" in user_role:
        file_id = request.POST.get('param0')
        index = request.POST.get('param1')
        model_by_column = request.POST.get('param2')
        #redirect_to_next_page = request.POST.get('param3')
        element_tag_id_index = request.POST.get('element_tag_id_index')
        file_table=ModelByColumn(model_by_column)
        elem_i_id ="table_span_"+model_by_column+"_"+file_id
        try:
            f_table_audit_v1 = file_table.objects.get(file_id=str(file_id),file_index=str(index))

            if f_table_audit_v1.file_removed == True:
                file_removed_path=f_table_audit_v1.file_save
                if os.path.exists(file_removed_path.path):
                    os.remove(file_removed_path.path)

                f_table_audit_v1.delete()
                return JsonResponse({'status': 'success',
                                    'title': 'Le fichier a été déplacé vers la corbeille',
                                    "element_tag_id_index":element_tag_id_index,})
            else:
                try:
                    Activity_table_object = TableData001.objects.get(cell_id=str(file_id)) #get_object_or_404(TableData001,cell_id=str(file_id))
                    if  Activity_table_object:
                        if getattr(Activity_table_object,"be"):
                            Activity_table="Bureau d'étude"
                        elif getattr(Activity_table_object,"ai"):
                            Activity_table="Agent immobilier"
                            
                    f_table_audit_v1.file_removed = True
                    #f_table_audit_v1.file_removed_date = 
                    f_table_audit_v1.file_removed_user_email = user_.email
                    f_table_audit_v1.file_removed_user_FN = user_.first_name
                    f_table_audit_v1.file_removed_user_LN = user_.last_name
                    f_table_audit_v1.file_removed_date=timezone.now()
                    f_table_audit_v1.save(update_fields=['file_removed', 'file_removed_user_email', 'file_removed_user_FN', 'file_removed_user_LN','file_removed_date' ])
                    model_by_column2= model_by_column
                    if model_by_column == "vt":
                        model_by_column="Visite technique"
                        model_by_column2= "VT"
                        elem_i_id ="table_span_"+model_by_column2+"_"+file_id
                    Activities_audit.objects.create(Activity_id=generate_random_string(10),
                                                                Activity_user=f"{user_.last_name} {user_.first_name}",
                                                                Activity_user_email=user_.email,
                                                                Activity_table=Activity_table,
                                                                Activity_project_id=str(file_id),
                                                                Activity_before=getattr(f_table_audit_v1,"file_name"),
                                                                Activity_after = model_by_column ,
                                                                Activity_delete=True)
                    return JsonResponse({'status': 'success',
                                    'title': 'Le fichier a été supprimé',
                                    "element_tag_id_index":element_tag_id_index,
                                    "id":file_id ,
                                    "index":index ,
                                    "column":model_by_column2,
                                    "elem_i_id":elem_i_id,})
                except:
                    pass
        except file_table.DoesNotExist:
            pass
        
    
        return JsonResponse({'status': 'error'})



@login_required
def agent_immo(request):

    data = TableData001.objects.filter(ai=True).order_by('-creation_time')
    
    col_count = data.count()
    # Get unique column names from the TableData model
    column_names = TableData001._meta.get_fields()
    datafiles_VT = file_table_vt.objects.all()
    datafiles_AuditV1 = file_table_auditV1.objects.all()
    datafiles_AuditV2 = file_table_auditV2.objects.all()
    datafiles_AuditV3 = file_table_auditV3.objects.all()
    datafiles_AuditFinal = file_table_auditFinal.objects.all()
    message_box_01 = message_box_1.objects.all()
    return render(request, 'html/agentimmo.html', { 'data': data ,
                                                  'col_count':col_count ,
                                                  'column_names': column_names,
                                                  'datafiles_VT': datafiles_VT ,
                                                  'datafiles_AuditFinal':datafiles_AuditFinal,
                                                  'message_box_1':message_box_01})
@login_required
def add_files_to_project_1(request,project_id,file_input_name,model_name):
    user_ =request.user
    if "ai" in user_.role:
        file_table = ModelByColumn(model_name)
        for file in request.FILES.getlist(file_input_name):
            format_file=file.name.split(".")[1]
            if format_file in ['jpg','png','jpeg','heic']:
                format_file="image"
            if format_file in ['doc','docx']:
                format_file="word"
            if format_file in ['xls','xlsm']:
                format_file="excel" 
            if not file_table.objects.filter(file_id=project_id, file_format=format_file,file_name=file.name):
                file_table.objects.create(
                        file_id = project_id,
                        file_name = file.name,
                        file_save = file,
                        file_format =format_file
                    )


@login_required
def agent_immo_f(request):
    acc_state=False
    user_ = request.user
    if "ai" in user_.role:
        
        if request.method == 'POST':
            
            project_id=generate_random_string(10)
            while TableData001.objects.filter(cell_id__contains = project_id,ai=True):
                project_id = generate_random_string(10)
                
            #myButton = request.POST.get("")
            firstname = request.POST.get("firstname")
            lastname = request.POST.get("lastname")
            address = request.POST.get("address")
            num = request.POST.get("num")
            agent = user_.com_name
            email = request.POST.get("email")
            TableData001.objects.create(cell_id=project_id,firstname=firstname,lastname=lastname,address=address,num=num,email=email,agent=agent,ai=True)
            acc_state=True
            column_name_type="VT"   
            if request.FILES.getlist("vt"):
                add_files_to_project_1(request,project_id,"vt",column_name_type)
                
            column_name_type="AdA"   
            if request.FILES.getlist("AdA"):
                add_files_to_project_1(request,project_id,"AdA",column_name_type)
            column_name_type="comm"   
            
            #if request.FILES.getlist("comm"):
                #add_files_to_project_1(request,project_id,"comm",column_name_type)
                
   
                        
    return render(request, 'html/agentimmoF.html',{"acc_state":acc_state} )


@login_required
def BE_home_page(request):
    user_=request.user
    data = TableData001.objects.all()
    activities_audit = Activities_be.objects.order_by("-Activity_date")[:6]
    today = timezone.now()
    first_day_of_year = date(today.year, 1, 1)
    first_day_of_month = date(today.year, today.month, 1)

    client_count_added_today = TableData001.objects.filter(creation_time__date=today,bureau_d_etude=user_.com_name).count()
    client_count_added_month = TableData001.objects.filter(creation_time__range=(first_day_of_month, timezone.now()),bureau_d_etude=user_.com_name).count()
    client_count_added_year= TableData001.objects.filter(creation_time__range=(first_day_of_year, timezone.now()),bureau_d_etude=user_.com_name).count()

    client_count_envoye_today = TableData001.objects.filter(Envoye_time__date=today,bureau_d_etude=user_.com_name).count()
    client_count_envoye_month = TableData001.objects.filter(Envoye_time__range=(first_day_of_month, timezone.now()),bureau_d_etude=user_.com_name).count()
    client_count_envoye_year= TableData001.objects.filter(Envoye_time__range=(first_day_of_year, timezone.now()),bureau_d_etude=user_.com_name).count()
    
    client_count_fini_today = TableData001.objects.filter(fini_time__date=today,bureau_d_etude=user_.com_name).count()
    client_count_fini_month = TableData001.objects.filter(fini_time__range=(first_day_of_month, timezone.now()),bureau_d_etude=user_.com_name).count()
    client_count_fini_year= TableData001.objects.filter(fini_time__range=(first_day_of_year, timezone.now()),bureau_d_etude=user_.com_name).count()



    return render(request, 'html/BE_home_page.html', {'data':data,
                                                        'activities_audit':activities_audit,
                                                        'client_count_added_today':client_count_added_today,
                                                        'client_count_added_month':client_count_added_month,
                                                        'client_count_added_year':client_count_added_year,
                                                        
                                                        'client_count_envoye_today':client_count_envoye_today,
                                                        'client_count_envoye_month':client_count_envoye_month,
                                                        'client_count_envoye_year':client_count_envoye_year,
                                                        
                                                        'client_count_fini_today':client_count_fini_today,
                                                        'client_count_fini_month':client_count_fini_month,
                                                        'client_count_fini_year':client_count_fini_year})

@login_required
def table_view_3(request):
    #Activity_id,Activity_user,Activity_table,Activity_in,Activity_before,Activity_after
    user_=request.user
    
    if request.method == 'POST':
        #return redirect("/Accueil/")
        if "be" in user_.role:
        
            file_uploaded_state=False
            files_date_for_response=[]
            button_edit_data_on_table=request.POST.get("cellId_new")
            file_Project_id =request.POST.get("cellId_new")
            
            if button_edit_data_on_table:

                try:
                    obj_by_id = get_object_or_404(TableData001,cell_id=str(button_edit_data_on_table))  # TableData001.objects.get(cell_id=str(button_edit_data_on_table))

                    table_index=[   {'column_name':'firstname','name2':'Prénom'},
                                    {'column_name':'lastname', 'name2':'Nom'},
                                    {'column_name':'address',  'name2':'Adresse'},
                                    {'column_name':'email',    'name2':'Email'},
                                    {'column_name':'num',      'name2':'N° de tél'}
                                    
                                    #{'column_name':'precaite', 'name2':'Précarite'}
                                    
                                    #{'column_name':'etat',     'name2':'État'},
                                    ]
                    
                    

                    Activity_table="Bureau d'étude"
                    
                    for l in table_index :
                        name = l['column_name']
                        name2 = l['name2']
                        re_page = False

                        Activity_before = str(getattr(obj_by_id, name) )+ ""
                        Activity_after = str(request.POST.get(f"table_{name}"))+ ""
                        if Activity_before != Activity_after:
                            if name == 'etat':
                                re_page=True
                            Activities_be.objects.create(Activity_id=generate_random_string(10),
                                                            Activity_user=f"{user_.last_name} {user_.first_name}",
                                                            Activity_user_email=user_.email,
                                                            Activity_table=Activity_table,
                                                            Activity_project_id=str(button_edit_data_on_table),
                                                            Activity_name=name2,
                                                            Activity_before=Activity_before,
                                                            Activity_after=Activity_after,
                                                            Activity_edit=True)
                            
                        setattr(obj_by_id, name, Activity_after)
                        obj_by_id.save()
                    
                    #if request.POST.get(f"table_etat") == "Envoye" and not obj_by_id.Envoye_time_checker :

                except TableData001.DoesNotExist:
                    pass
                
            VT_files_added = 0 
            auditV1_files_added = 0
            auditV2_files_added = 0
            auditV3_files_added = 0
            auditFinal_files_added = 0
            
            column_name_types = ["VT", "auditV1", "auditV2", "auditV3", "auditFinal"]
            for column_name_type in column_name_types:
                files_uploaded = request.FILES.getlist(f"table_{column_name_type}[]")
                if files_uploaded:

                    file_uploaded_state = True
                    
                    l1=f"table_{column_name_type}[]"
                    file_table = ModelByColumn(column_name_type)
                    files_date_for_response=[]
                    I_icon_class=""

                    for file in request.FILES.getlist(l1, []):
                        if column_name_type == "VT":
                            VT_files_added +=1
                            
                        if column_name_type == "auditV1":
                            auditV1_files_added +=1
                           
                        if column_name_type == "auditV2":
                            auditV2_files_added +=1
                           
                        if column_name_type == "auditV3":
                            auditV3_files_added +=1
                            
                        if column_name_type == "auditFinal":
                            auditFinal_files_added +=1
                            
                        format_file=file.name.split(".")[1]
                        if format_file not in ['jpg','png','jpeg','heic','doc','docx','xls','xlsm','pdf','pz2']:
                            I_icon_class="ri-file-line"
                            
                        if format_file in ['jpg','png','jpeg','heic']:
                            format_file="image"
                            I_icon_class="bi bi-file-earmark-image"
                        if format_file in ['doc','docx']:
                            format_file="word"
                            I_icon_class="bi bi-file-earmark-word"
                        if format_file in ['xls','xlsm']:
                            format_file="excel" 
                            I_icon_class="ri-file-excel-2-line"
                        if format_file in ['pdf']:
                            format_file="pdf" 
                            I_icon_class="bi bi-file-earmark-pdf"
                        if format_file in ['pz2']:
                            format_file="pz2" 
                            I_icon_class="bi bi-file-ppt"
                        
                        
                        
                        if not file_table.objects.filter(file_id=button_edit_data_on_table, file_name=file.name,file_removed=False):
                            Activities_be.objects.create(
                                    Activity_id=generate_random_string(10),
                                    Activity_user = f"{request.user.last_name} {request.user.first_name}",
                                    Activity_user_email = request.user.email,
                                    Activity_table=Activity_table,
                                    Activity_project_id = button_edit_data_on_table,
                                    Activity_before =f"le fichier {file.name}" ,
                                    Activity_after = column_name_type ,
                                    Activity_add=True
                                )
                            file_table.objects.create(
                                    file_id = button_edit_data_on_table,
                                    file_name = file.name,
                                    file_save = file,
                                    file_format =format_file
                                )
                            column_name_type_new = column_name_type
                            if column_name_type == "VT":
                                column_name_type_new="vt"
                            files_date_for_response.append({'file_id':button_edit_data_on_table,
                                                        'file_save_url':"file.name",
                                                        'file_format':format_file,
                                                        'file_name':file.name,
                                                        'file_index':"",
                                                        'column':column_name_type_new,
                                                        'I_icon_class':I_icon_class,})

                        
            files_added_list=[{'files_added':VT_files_added,'file_id':file_Project_id,'col_n':"VT"},
                              {'files_added':auditV1_files_added,'file_id':file_Project_id,'col_n':"auditV1"},
                              {'files_added':auditV2_files_added,'file_id':file_Project_id,'col_n':"auditV2"},
                              {'files_added':auditV3_files_added,'file_id':file_Project_id,'col_n':"auditV3"},
                              {'files_added':auditFinal_files_added,'file_id':file_Project_id,'col_n':"auditFinal"}]
            response_date={
                're_page':re_page,
                'cellId_new':button_edit_data_on_table,
                'file_uploaded_state':file_uploaded_state,
                'files_date_for_response':files_date_for_response,
                'files_added_list':files_added_list,
            }
            
            return JsonResponse(response_date)
    return JsonResponse({'status': 'error'})

@login_required
def CBFCS(request):
    user_=request.user
    
    if request.method == 'POST':
        if "be" in user_.role:
            cell_id = request.POST.get("param0")
            try:
                    obj_by_id = get_object_or_404(TableData001,cell_id=str(cell_id))  # TableData001.objects.get(cell_id=str(button_edit_data_on_table))


                    name = "etat"
                    name2 = "État"
                    re_page = False

                        
                    Activity_before = "Fini"
                    Activity_after = "A modifier"
                    if Activity_before != Activity_after:

                        re_page=True
                        Activities_be.objects.create(Activity_id=generate_random_string(10),
                                                        Activity_user=f"{user_.last_name} {user_.first_name}",
                                                        Activity_user_email=user_.email,
                                                        Activity_project_id=str(cell_id),
                                                        Activity_name=name2,
                                                        Activity_before=Activity_before,
                                                        Activity_after=Activity_after,
                                                        Activity_edit=True)
                        
                        
                    setattr(obj_by_id, name, Activity_after)
                    setattr(obj_by_id, 'Modification_Faite_time_checker', True)
                    setattr(obj_by_id, 'Modification_Faite_time_checker_2', True)
                    obj_by_id.save()
                    
                    response_date={
                            're_page':re_page,
                        }
                        
            except:
                pass
            
                
            return JsonResponse(response_date)
    return JsonResponse({'status': 'error'})
@login_required
def BE_Page(request):
    
    user_ = request.user
    bureau_d_etude = user_.com_name
    data = TableData001.objects.filter(be=True,bureau_d_etude=bureau_d_etude).order_by('-creation_time')
    table_index=[{'index':1,'state':"A realiser",'state_2':"A realiser ;  En cours ; "},
                 {'index':2,'state':"Fini",'state_2':"Fini"},
                 {'index':3,'state':"A modifier",'state_2':"A modifier ; Modification Faite"}]
    
    col_count = data.count()
    # Get unique column names from the TableData model
    column_names = TableData001._meta.get_fields()
    datafiles_VT = file_table_vt.objects.filter(file_removed=False)
    countfiles_VT = file_table_vt.objects.filter(file_removed=False).count()
    datafiles_AuditV1 = file_table_auditV1.objects.filter(file_removed=False)
    datafiles_AuditV2 = file_table_auditV2.objects.filter(file_removed=False)
    datafiles_AuditV3 = file_table_auditV3.objects.filter(file_removed=False)
    datafiles_AuditFinal = file_table_auditFinal.objects.filter(file_removed=False)
    message_box_01 = message_box_1.objects.all()
    
    msg_box_tagged=[user_.email,user_.first_name,user_.last_name]
    
    return render(request, 'html/BE_page.html', { 'data': data ,
                                                 'table_index':table_index,
                                                  'col_count':col_count ,
                                                  'column_names': column_names,
                                                  'datafiles_VT': datafiles_VT ,
                                                  'countfiles_VT': countfiles_VT ,
                                                  'datafiles_AuditV1': datafiles_AuditV1 ,
                                                  'datafiles_AuditV2': datafiles_AuditV2 ,
                                                  'datafiles_AuditV3': datafiles_AuditV3 ,
                                                  'datafiles_AuditFinal':datafiles_AuditFinal,
                                                  'message_box_1':message_box_01,
                                                  'redirect_next_page':redirect_page_be,
                                                  'msg_box_tagged':msg_box_tagged,})
@login_required
def BE_Page_f(request):
    user_= request.user
    user_role= user_.role
    if  "be" in user_role:
        
        if request.method == 'POST' :
            cell_id =generate_random_string(10)
            while TableData001.objects.filter(cell_id=cell_id).exists():
                cell_id =generate_random_string(10)
            
            firstname = request.POST.get('firstname')
            lastname = request.POST.get('lastname')
            address = request.POST.get('address')
            email = request.POST.get('email')
            num = request.POST.get('num')
            precaite = request.POST.get('precaite')
            bureau_d_etude = user_.com_name
            
            TableData001.objects.create(cell_id=cell_id,
                                        firstname=firstname,
                                        lastname=lastname,
                                        address=address,
                                        email=email,
                                        num=num,
                                        precaite=precaite,
                                        be=True,
                                        bureau_d_etude=bureau_d_etude)
            
            kizeo_model.objects.create(kizeo_id=cell_id,)

            
            Activities_be.objects.create(
                            Activity_id=generate_random_string(10),
                            Activity_user = f"{request.user.last_name} {request.user.first_name}",
                            Activity_user_email = request.user.email,
                            Activity_project_id = cell_id,
                            Activity_add_client=True
                        )
            
            for file in  request.FILES.getlist('vt'):

                file_table_vt.objects.create(file_id=cell_id,
                                            file_name=file.name,
                                            file_save=file,
                                            file_format=file.name.split(".")[1]
                                            )
                

                
                
        return render(request, 'html/BEform.html', )



@login_required
def VT_Page(request):

    table_index=[{'index':1,'state':"En cours"},
                 {'index':2,'state':"Fini"}]
    data = TableData001.objects.all()
    return render(request, 'html/VTPage.html', { 'data': data,'table_index':table_index })
@login_required
def VT_Page_edit_state(request):
    user_ = request.user
    user_role= user_.role
    if  "VT" in user_role:
        param_value_id = request.GET.get('param0')
        etat_vt_value = request.GET.get('param1')
        
        try:
            object = TableData001.objects.get(cell_id=str(param_value_id))
            object.etat_vt =  etat_vt_value
            object.save(update_fields=['etat_vt'])

        except TableData001.DoesNotExist:
            pass
    
    
    return redirect(VT)
@login_required
def create_acc_ai(request):
    acc_state = False
    user_=request.user
    if "a1" in user_.role:
        if request.method == 'POST': 
            if request.POST.get('add_com') == "submit":
                nom = request.POST.get('nom')
                AI_or_AGENT_id = generate_random_string(10)
                while AI_or_AGENT.objects.filter(AI_or_AGENT_id = AI_or_AGENT_id):
                    AI_or_AGENT_id = generate_random_string(10)
                AI_or_AGENT.objects.create(AI_or_AGENT_id=AI_or_AGENT_id,
                                        comp_name=nom,
                                        ai=True)
                Activities_audit.objects.create(
                        Activity_id=generate_random_string(10),
                        Activity_user = f"{user_.last_name} {user_.first_name}",
                        Activity_user_email = user_.email,
                        Activity_before = f"un agent immobilier ( {nom} )",
                        Activity_add_2=True
                    )
                return redirect("/create_account_for_ai/")
                
            if request.POST.get('create_acc_button') == "submit":
                
                firstname = request.POST.get('firstname')
                lastname = request.POST.get('lastname')
                email = request.POST.get('email')
                Num = int(request.POST.get('num'))
                Agent = request.POST.get('agent')
                acc_for = "; ai ;"
                user_id = generate_random_string(10)
                password = generate_random_string(12)
                profile_pic="uploads/default_user_avatar.png"
                while USER.objects.filter(user_id__contains = user_id):
                    user_id = generate_random_string(10)
                    
                user_checker=USER.objects.filter(email__icontains = email).exists()
                if not user_checker:
                    
                    subject = 'Votre compte a été créé avec succès'
                    message = f'Email: {email} Mot de passe: {password} '
                    from_email = 'Night'
                    recipient_list = [email]
                    send_mail(subject, message, from_email, recipient_list) 
                    
                    USER.objects.create_user(first_name=firstname,
                                        last_name=lastname,
                                        email=email,
                                        num=Num,
                                        role=acc_for, 
                                        password=password,
                                        ai=True,
                                        com_name=Agent,
                                        profile_pic=profile_pic)
                    acc_state = True
                    Activities_audit.objects.create(
                        Activity_id=generate_random_string(10),
                        Activity_user = f"{user_.last_name} {user_.first_name}",
                        Activity_user_email = user_.email,
                        Activity_before = f"un utilisateur ' Agent immobilier ' ( {email} - {lastname} {firstname} )",
                        Activity_add_2=True
                    )
                
                  
    comp = AI_or_AGENT.objects.all()
    
    return render(request, 'html/create_acc_ai.html',{'acc_state':acc_state,
                                                      'comp':comp})
    
@login_required
def create_acc_be(request):
    acc_state = False
    user_=request.user
    if "a1" in user_.role:
        if request.method == 'POST': 
            if request.POST.get('add_com') == "submit":
                nom = request.POST.get('nom')
                AI_or_AGENT_id = generate_random_string(10)
                while AI_or_AGENT.objects.filter(AI_or_AGENT_id = AI_or_AGENT_id):
                    AI_or_AGENT_id = generate_random_string(10)
                AI_or_AGENT.objects.create(AI_or_AGENT_id=AI_or_AGENT_id,
                                        comp_name=nom,
                                        be=True)
                Activities_audit.objects.create(
                        Activity_id=generate_random_string(10),
                        Activity_user = f"{user_.last_name} {user_.first_name}",
                        Activity_user_email = user_.email,
                        Activity_before = f"un bureau d'études ( {nom} )",
                        Activity_add_2=True
                    )
                return redirect("/create_account_for_be/")
            
            if request.POST.get('create_acc_button') == "submit":
                
                firstname = request.POST.get('firstname')
                lastname = request.POST.get('lastname')
                email = request.POST.get('email')
                Num = int(request.POST.get('num'))
                Agent = request.POST.get('agent')
                acc_for = "; be ;"
                
                Bureau_etude = request.POST.get('Bureau_etude')
                
                user_id = generate_random_string(10)
                
                password = generate_random_string(12)
                profile_pic="uploads/default_user_avatar.png"
                while USER.objects.filter(user_id__contains = user_id):
                    user_id = generate_random_string(10)
                    
                user_checker=USER.objects.filter(email__icontains = email).exists()
                if not user_checker:
                    
                    subject = 'Votre compte a été créé avec succès'
                    message = f'Email: {email}  ||   Mot de passe: {password} '
                    from_email = f'{FROM_EMAIL}'
                    recipient_list = [email]
                    send_mail(subject, message, from_email, recipient_list) 
                    
                    USER.objects.create_user(first_name=firstname,
                                        last_name=lastname,
                                        email=email,
                                        num=Num,
                                        role=acc_for, 
                                        password=password,
                                        be=True,
                                        com_name=Bureau_etude,
                                        profile_pic=profile_pic)
                    acc_state = True
                    Activities_audit.objects.create(
                        Activity_id=generate_random_string(10),
                        Activity_user = f"{user_.last_name} {user_.first_name}",
                        Activity_user_email = user_.email,
                        Activity_before = f"un utilisateur ' Bureau d'études ' ( {email} - {lastname} {firstname} )",
                        Activity_add_2=True
                    )


    comp = AI_or_AGENT.objects.all()
    return render(request, 'html/create_acc_be.html',{'acc_state':acc_state,
                                                      'comp':comp})
    
@login_required
def create_acc_auditeur(request):
    acc_state = False
    user_=request.user
    if "a1" in user_.role:
        if request.method == 'POST': 
            if request.POST.get('add_com') == "submit":
                nom = request.POST.get('nom')
                AI_or_AGENT_id = generate_random_string(10)
                while AI_or_AGENT.objects.filter(AI_or_AGENT_id = AI_or_AGENT_id):
                    AI_or_AGENT_id = generate_random_string(10)
                AI_or_AGENT.objects.create(AI_or_AGENT_id=AI_or_AGENT_id,
                                        comp_name=nom,
                                        ai=True)
                Activities_audit.objects.create(
                        Activity_id=generate_random_string(10),
                        Activity_user = f"{user_.last_name} {user_.first_name}",
                        Activity_user_email = user_.email,
                        Activity_before = f"un auditeur ( {nom} )",
                        Activity_add_2=True
                    )
                return redirect("/create_account_for_ai/")
                
            if request.POST.get('create_acc_button') == "submit":
                
                firstname = request.POST.get('firstname')
                lastname = request.POST.get('lastname')
                email = request.POST.get('email')
                Num = int(request.POST.get('num'))
                Agent = request.POST.get('agent')
                acc_for = "; auditeur ;"
                user_id = generate_random_string(10)
                password = generate_random_string(12)
                profile_pic="uploads/default_user_avatar.png"
                while USER.objects.filter(user_id__contains = user_id):
                    user_id = generate_random_string(10)
                    
                user_checker=USER.objects.filter(email__icontains = email).exists()
                if not user_checker:
                    
                    subject = 'Votre compte a été créé avec succès'
                    message = f'Email: {email} Mot de passe: {password} '
                    from_email = 'Night'
                    recipient_list = [email]
                    send_mail(subject, message, from_email, recipient_list) 
                    
                    USER.objects.create_user(first_name=firstname,
                                        last_name=lastname,
                                        email=email,
                                        num=Num,
                                        role=acc_for, 
                                        password=password,
                                        com_name=Agent,
                                        auditeur=True,
                                        profile_pic=profile_pic)
                    acc_state = True
                    Activities_audit.objects.create(
                        Activity_id=generate_random_string(10),
                        Activity_user = f"{user_.last_name} {user_.first_name}",
                        Activity_user_email = user_.email,
                        Activity_before = f"un utilisateur ' Auditeur ' ( {email} - {lastname} {firstname} )",
                        Activity_add_2=True
                    )

                    
                       
    comp = AI_or_AGENT.objects.all()
    
    return render(request, 'html/create_acc_auditeur.html',{'acc_state':acc_state,
                                                      'comp':comp})
  
       
    
@login_required
def corbeille(request):
    redirect_page="/historique_des_fichiers/"
    datafiles_VT = file_table_vt.objects.filter(file_removed=True)
    datafiles_AuditV1 = file_table_auditV1.objects.filter(file_removed=True)
    datafiles_AuditV2 = file_table_auditV2.objects.filter(file_removed=True)
    datafiles_AuditV3 = file_table_auditV3.objects.filter(file_removed=True)
    datafiles_AuditFinal = file_table_auditFinal.objects.filter(file_removed=True)
    #model_type=['vt','auditV1','auditV2','auditV3','auditFinal']
    model_type=[{'index':1,'model':datafiles_VT,'name':"vt"},
                {'index':2,'model':datafiles_AuditV1,'name':"auditV1"},
                {'index':3,'model':datafiles_AuditV2,'name':"auditV2"},
                {'index':4,'model':datafiles_AuditV3,'name':"auditV3"},
                {'index':5,'model':datafiles_AuditFinal,'name':"auditFinal"}]
    return render(request, 'html/corbeille.html',{'datafiles_VT': datafiles_VT ,
                                                  'datafiles_AuditV1': datafiles_AuditV1 ,
                                                  'datafiles_AuditV2': datafiles_AuditV2 ,
                                                  'datafiles_AuditV3': datafiles_AuditV3 ,
                                                  'datafiles_AuditFinal':datafiles_AuditFinal,
                                                  'redirect_next_page':redirect_page,
                                                  'model_type':model_type})





## <a href="{% url 'download_media_folder' %}">Download Media Folder</a>

@login_required
def download_media_folder(request):
    date_time_now=timezone.now()
    downloaded_folders=date_time_now.strftime("%Y-%m-%d_%H-%M-%S")
    
    user_ = request.user
    if "backup_server_ROLE_for_download_the_backUP" in user_.role :
        other_folder = BASE_DIR
        dump_file_path = os.path.join(other_folder, 'backup.json')
        cwd_ = BASE_DIR
        # Run the dumpdata command
        dumpdata_cmd = "python manage.py dumpdata --indent 2 > backup.json"
        subprocess.run(dumpdata_cmd, shell=True, cwd=cwd_)

        # Read the dumped data from the file
        with open(dump_file_path, 'r') as f:
            data = f.read()

        # Prepare the response with the dumped data
        response = HttpResponse(data, content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="exported_data.json"'
        # Replace 'your_media_folder' with the relative path to your media folder


        media_folder = media_path =  BASE_DIR / "media"
        other_folder = BASE_DIR
            # Get the absolute paths to the folders
        media_path = os.path.join('media', media_folder)

        # Ensure that both folders exist
        if os.path.exists(media_path) and os.path.exists(other_folder):
            # Create a temporary directory to store the zip file
            temp_dir = tempfile.mkdtemp()

            # Create a zip file containing the contents of both folders
            zip_filename = f'{downloaded_folders}.zip'
            zip_filepath = os.path.join(temp_dir, zip_filename)

            with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # Add files from the media folder
                for root, dirs, files in os.walk(media_path):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, media_path)
                        zip_file.write(file_path, arcname)

                # Add the JSON file from the other folder
                json_file_path = os.path.join(other_folder, 'backup.json')  # Replace with the actual filename
                arcname = os.path.basename(json_file_path)
                zip_file.write(json_file_path, arcname)

            # Prepare the response with the zip file
            with open(zip_filepath, 'rb') as zip_file:
                response = HttpResponse(zip_file.read(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'

            # Clean up the temporary directory and zip file
            shutil.rmtree(temp_dir)
            
            os.remove(dump_file_path)
            return response




@login_required
def Activities(request):
    activities_audit = Activities_audit.objects.order_by("-Activity_date")
    
    return render(request, 'html/activities.html',{'activities_audit': activities_audit})
    
@login_required
def update_xlsx_template(request):
    # Load your XLSX template
    
    template_path = 'ERapp\static\Kizeo.xlsx'  # Provide the path to your template file
    workbook = openpyxl.load_workbook(template_path)
    
    ws1 = workbook["Données"]
    worksheet3 =  workbook["Garde"]
    worksheet4 =  workbook["Site"]
    #for i in range(2,17):
        #ws1[f"B{i}"].value ="night code"
    
    ws1["B2"].value ="night code"
    ws1["B4"].value ="derb seltan"
    
    
    ############################# WORK SHEET 2
    ############################# WORK SHEET 2
    ############################# WORK SHEET 2
    ############################# WORK SHEET 2
    ############################# WORK SHEET 2
    worksheet = workbook["Métré"]  # Select the first sheet or specify the sheet name
    # Fetch data from your Django model (Assuming you have a queryset)
    obj = MyModel.objects.get(text_field="t1")

    


    
    for i in range(3,17):
        
        # Find and update text cells
        text_cell = worksheet[f'f{i}']  # Example: Replace 'A1' with the actual cell containing text
        text_cell.value = "non" #obj.text_field

        for c_name in ["A","B","C"]:
            # Find and update image cells (assuming image_field is a Django ImageField)
            image_cell = worksheet[f'{c_name}{i}']  # Example: Replace 'B1' with the actual cell containing the image
            image_cell.value = None
            #image_cell.hyperlink = None
            image_list = [obj.image_field.path,obj.image_field_2.path,obj.image_field_3.path]
            image_path = random.choice(image_list)
            
            
            # Load the image
            img = Image(image_path)

            # Calculate the image size to fit the cell
            cell_width = worksheet.column_dimensions[get_column_letter(image_cell.column)].width
            cell_height = worksheet.row_dimensions[image_cell.row].height
            img.width =  199.68 #cell_width
            img.height = 149.76# cell_height
            
            worksheet.column_dimensions[c_name].width = 29.86
            #worksheet.row_dimensions['4'].height = 127
            
            # Add the new image to the cell
            worksheet.add_image(img, image_cell.coordinate)
        
    ############################# END WORK SHEET 2
    ############################# END WORK SHEET 2
    ############################# END WORK SHEET 2
    ############################# END WORK SHEET 2

    
    
    ############################################### sheet 3 image
    image_cell2 = worksheet3['B6']  # Example: Replace 'B1' with the actual cell containing the image
    image_cell2.value = None
    #image_cell2.hyperlink = None
    image_path2 = obj.image_field.path

    # Load the image
    img2 = Image(image_path2)

    # Calculate the image size to fit the cell
    img2.width =  371.52 #cell_width
    img2.height = 280.32# cell_height
    #worksheet3.column_dimensions['B6'].width = 29.86
    #worksheet.row_dimensions['4'].height = 127
    # Add the new image to the cell
    worksheet3.add_image(img2, image_cell2.coordinate)
    
    
    
    
    ############################## header
    ############################## header
    ############################## header
    # Load the image
    img3 = Image(image_path2)
    # Calculate the image size to fit the cell
    img3.width =  50 #cell_width
    img3.height = 30# cell_height
    worksheet3.oddHeader.center.text = "&L &G &C Your Image Placeholder &R"
    ############################## header
    ############################## header
    ############################## header
        
    
    
    
    ######################################   WORK_SHEET_4
    ######################################   WORK_SHEET_4
    ######################################   WORK_SHEET_4
    # A38 E38 
    # A50 E50
    # B67
    
    image_cell_1 = worksheet4['A38']  # Example: Replace 'B1' with the actual cell containing the image
    image_cell_1.value = None
    #image_cell_1.hyperlink = None
    image_path1 = obj.image_field_2.path
    img_1 = Image(image_path1)
    img_1.width =  185.28 #cell_width
    img_1.height = 140.16 # cell_height
    worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    image_cell_1 = worksheet4['E38']  # Example: Replace 'B1' with the actual cell containing the image
    image_cell_1.value = None
    #image_cell_1.hyperlink = None
    # image url
    image_path1 = obj.image_field_3.path
    img_1 = Image(image_path1)
    img_1.width =  185.28 #cell_width
    img_1.height = 140.16 # cell_height
    worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    image_cell_1 = worksheet4['E50']  # Example: Replace 'B1' with the actual cell containing the image
    image_cell_1.value = None
    #image_cell_1.hyperlink = None
    image_path1 = obj.image_field_2.path
    img_1 = Image(image_path1)
    img_1.width =  185.28 #cell_width
    img_1.height = 140.16 # cell_height
    worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    image_cell_1 = worksheet4['A50']  # Example: Replace 'B1' with the actual cell containing the image
    image_cell_1.value = None
    #image_cell_1.hyperlink = None
    # image url
    image_path1 = obj.image_field_3.path
    img_1 = Image(image_path1)
    img_1.width =  185.28 #cell_width
    img_1.height = 140.16 # cell_height
    worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    image_cell_1 = worksheet4['B67']  # Example: Replace 'B1' with the actual cell containing the image
    image_cell_1.value = None
    #image_cell_1.hyperlink = None
    # image url
    image_path1 = obj.image_field_3.path
    img_1 = Image(image_path1)
    img_1.width =  185.28 #cell_width
    img_1.height = 140.16 # cell_height
    worksheet4.add_image(img_1, image_cell_1.coordinate)
    ######################################   END WORK_SHEET_4
    ######################################   END WORK_SHEET_4
    ######################################   END WORK_SHEET_4
    
    


    ###################################################### save process #######################################################
    # Save the updated XLSX file temporarily
    temp_xlsx_path = 'path_to_temp_xlsx.xlsx'  # Provide a temporary path on your server
    workbook.save(temp_xlsx_path)

    # Create a new instance of your other model
    updated_xlsx = UpdatedXLSXFile()

    # Assign the XLSX file to the FileField or ImageField of the new model instance
    updated_xlsx.xlsx_file.save('updated_template.xlsx', open(temp_xlsx_path, 'rb'))

    # Save the new model instance to persist the XLSX file
    updated_xlsx.save()

    return HttpResponse('XLSX file updated and saved to another model.')


class table_index_image_fgfg:
    def __init__(self,index,obj_image):
        self.index=index
        self.obj_image=obj_image
        
@login_required        
def Kizeo_form_page(request,client_id,page_number):
    user_ = request.user
    if kizeo_model.objects.filter(kizeo_id=client_id):
            pass
    else:
        kizeo_model.objects.create(kizeo_id=client_id)
    
    obj = kizeo_model.objects.get(kizeo_id=client_id)
    if "kize" in user_.role:
        if request.method == 'POST':
            
            myButton = request.POST.get("mybutton1")
            if myButton=="mybutton1":

                input_cell_name = request.POST.get("input_value")
                img_get= request.FILES[input_cell_name]
            
                if hasattr(obj, input_cell_name):
                    # Check if the field exists in the model
                    setattr(obj, input_cell_name, img_get)
                    obj.save()        
            
            next_page_index=""
            submit_to_Kizeo = request.POST.get("submit_to_Kizeo")
            list_inputs=[]
            table_index=[]
            if str(submit_to_Kizeo)=="1":
                next_page_index='2'
                list_inputs=[### Données Générales
                                        'latitude',
                                        'longitude',
                                        'altitude',
                                        'Donnees_Generales_Nom_d_intervenant',
                                        'Donnees_Generales_Date_de_visite',
                                        'Donnees_Generales_Adresse',
                                        'Donnees_Generales_Zip_Code',
                                        'Donnees_Generales_City',
                                        'Donnees_Generales_Annee_de_construction',
                                        'Donnees_Generales_Etat_d_occupation',
                                        'Donnees_Generales_Nom_client',
                                        'Donnees_Generales_Tel_client',
                                        'Donnees_Generales_Email',
                                        'Donnees_Generales_Horaire_d_occupation_des_lieux',
                                        'Donnees_Generales_Destination_du_lieu',
                                        'Donnees_Generales_Nombre_d_occupant',
                                        'Donnees_Generales_Nombre_de_niveau',
                                        'Donnees_Generales_Surface_TOTALE',
                                        'Donnees_Generales_Preuve_Surface',
                                        'Donnees_Generales_Surface_ajoute_depuis_moins_de_15_ans',
                                        'Donnees_Generales_Besoin_du_client_Chauffage',
                                        'Donnees_Generales_Besoin_du_client_Isolation',
                                        'Donnees_Generales_Scenario_souhaite_par_le_client',
                                        
                                        ### Façades
                                        'Facade_1_Orientation', 'Facade_1_Mitoyennete','Facade_1_Longueur','Facade_1_Hauteur','Facade_1_Surface',#'Facade_1_Photo_Principale',
                                        'Facade_2_Orientation', 'Facade_2_Mitoyennete','Facade_2_Longueur','Facade_2_Hauteur','Facade_2_Surface',#'Facade_2_Photo_Principale',
                                        'Facade_3_Orientation', 'Facade_3_Mitoyennete','Facade_3_Longueur','Facade_3_Hauteur','Facade_3_Surface',#'Facade_3_Photo_Principale',
                                        'Facade_4_Orientation', 'Facade_4_Mitoyennete','Facade_4_Longueur','Facade_4_Hauteur','Facade_4_Surface',#'Facade_4_Photo_Principale',
                            ]
                table_index=['Donnees_Generales_Preuve_Surface_Photo',
                            'Donnees_Generales_Factures',
                            'Facade_1_Photo_Principale',
                            'Facade_2_Photo_Principale',
                            'Facade_3_Photo_Principale',
                            'Facade_4_Photo_Principale',
                            ]
            if str(submit_to_Kizeo)=="2":
                next_page_index='3'
                list_inputs=[### Cauffage
                                        "Cauffage_systeme",
                                        "Cauffage_annee_de_mise_en_oeuvre",
                                        "Cauffage_type_de_regulation",
                                        "Cauffage_system_d_appoint",
                                        "Cauffage_commentaire" ,
                                        
                                        ### ECS
                                        "ECS_type",
                                        "ECS_system_d_appoint",
                                        "ECS_commentaire",
                                        
                                        ### Ventilation
                                        "Ventilation_type" ,
                                        
                                        ### Refroidissement
                                        "Refroidissement_type" ,
                                        "Refroidissement_commentaire" ,
                                        
                                        ### Compteur Electrique
                                        "Compteur_Electrique_Puissance_souscrite" ,
                                        "Compteur_Electrique_type" ,
                                        "Compteur_Electrique_commentaire" ,]
                table_index=['Cauffage_photo_systeme_de_production',
                            'Cauffage_photo_fiche_signaletique',
                            'Cauffage_photo_appoint',
                            'ECS_photo_appoint',
                            'Ventilation_photo_ventilation',
                            'Compteur_Electrique_photo_compteur',
                            ]
            if str(submit_to_Kizeo)=="3":
                next_page_index='4'
                list_inputs=[### Mur 1
                                        "Mur_1_Position", 
                                        "Mur_1_Composition",
                                        "Mur_1_Epaisseur_mur",
                                        "Mur_1_Isolation",
                                        "Mur_1_Epaisseur_isolant",
                                        "Mur_1_Date_d_isolation",
                                        "Mur_1_Preuve_d_isolation",
                                        ### Mur 2
                                        "Mur_2_Position", 
                                        "Mur_2_Composition",
                                        "Mur_2_Epaisseur_mur",
                                        "Mur_2_Isolation",
                                        "Mur_2_Epaisseur_isolant",
                                        "Mur_2_Date_d_isolation",
                                        "Mur_2_Preuve_d_isolation",
                                        
                                        ### Plancher bas 1
                                        "Plancher_bas_1_Position",
                                        "Plancher_bas_1_Composition" ,
                                        "Plancher_bas_1_Isolation" ,
                                        "Plancher_bas_1_Epaisseur_isolant" ,
                                        "Plancher_bas_1_Date_d_isolation" ,
                                        "Plancher_bas_1_Preuve_d_isolation" ,
                                        #"Plancher_bas_1_Photo_plancher_bas" ,
                                        ### Plancher bas 2
                                        "Plancher_bas_2_Position",
                                        "Plancher_bas_2_Composition" ,
                                        "Plancher_bas_2_Isolation" ,
                                        "Plancher_bas_2_Epaisseur_isolant" ,
                                        "Plancher_bas_2_Date_d_isolation" ,
                                        "Plancher_bas_2_Preuve_d_isolation" ,
                                        #"Plancher_bas_2_Photo_plancher_bas" ,
                                        
                                        ### Plancher haut 1
                                        "Plancher_Haut_1_Type",
                                        "Plancher_Haut_1_Composition",
                                        "Plancher_Haut_1_Surface",
                                        "Plancher_Haut_1_Isolation" ,
                                        "Plancher_Haut_1_Epaisseur_isolant" ,
                                        "Plancher_Haut_1_Date_d_isolation",
                                        "Plancher_Haut_1_Preuve_d_isolation" ,
                                        #"Plancher_Haut_1_Photo_plancher_bas" ,
                                        ### Plancher haut 2
                                        "Plancher_Haut_2_Type",
                                        "Plancher_Haut_2_Composition",
                                        "Plancher_Haut_2_Surface",
                                        "Plancher_Haut_2_Isolation" ,
                                        "Plancher_Haut_2_Epaisseur_isolant" ,
                                        "Plancher_Haut_2_Date_d_isolation",
                                        "Plancher_Haut_2_Preuve_d_isolation" ,
                                        #"Plancher_Haut_2_Photo_plancher_bas" ,
                                        ]
                table_index=['Mur_1_Photo_mur',
                            'Mur_2_Photo_mur',
                            'Plancher_bas_1_Photo_plancher_bas',
                            'Plancher_bas_2_Photo_plancher_bas',
                            'Plancher_Haut_1_Photo_plancher_bas',
                            'Plancher_Haut_2_Photo_plancher_bas',
                    ]
            if str(submit_to_Kizeo)=="4":
                next_page_index='5'
                list_inputs=[### Fenetre type 1
                                        "Fenetre_type_1_Menuiserie" ,
                                        "Fenetre_type_1_Materiaux",
                                        "Fenetre_type_1_Type_de_vitrage" ,
                                        "Fenetre_type_1_Volets" ,
                                        "Fenetre_type_1_Nombre",
                                        #"Fenetre_type_1_Photo" ,
                                        ### Fenetre type 2
                                        "Fenetre_type_2_Menuiserie" ,
                                        "Fenetre_type_2_Materiaux",
                                        "Fenetre_type_2_Type_de_vitrage" ,
                                        "Fenetre_type_2_Volets" ,
                                        "Fenetre_type_2_Nombre",
                                        #"Fenetre_type_2_Photo" ,
                                        
                                        ### Porte 1
                                        "Porte_1_Materiaux",
                                        "Porte_1_Type_porte",
                                        "Porte_1_Nombre",
                                        ### Porte 2
                                        "Porte_2_Materiaux",
                                        "Porte_2_Type_porte",
                                        "Porte_2_Nombre",]
                table_index=['Fenetre_type_1_Photo',
                            'Fenetre_type_2_Photo',
                            'Porte_1_Photo_porte',
                            'Porte_2_Photo_porte',
                    ]
                
            if str(submit_to_Kizeo)=="5":
                next_page_index='0'
                list_inputs=[   ###  Saisie par pièce
                            'Saisie_par_piece_Surface_Mesuree',
                            ]
                table_index=[
                            'Saisie_par_piece_Signature_intervenant',
                            'Saisie_par_piece_Signature_client'
                    ]
            for lis_inp in list_inputs:
                    input_get= request.POST.get(lis_inp)
                    if input_get:
                        setattr(obj, lis_inp, input_get)
                        obj.save()
                        
            for lis in table_index:
                    img_get= request.FILES.get(lis)
                    if img_get:
                        if hasattr(obj, lis):
                            if getattr(obj,lis):
                                if len(getattr(obj,lis)) > 0 :
                                    os.remove(getattr(obj,lis).path)
                            # Check if the field exists in the model
                            setattr(obj, lis, img_get)
                            obj.save()  
        
        
            return redirect(f"/formK/{client_id}/{next_page_index}")
    Pieces_index_add = kizeo_model_Pieces.objects.filter(kizeo_id=client_id).aggregate(Max('Pieces_index'))['Pieces_index__max']
    if not Pieces_index_add:
        Pieces_index_add = 1
    else: 
        Pieces_index_add+=1
        
    Pieces_index_add = kizeo_model_Pieces.objects.aggregate(Max('pk'))['pk__max']
    Pieces_index_add+=1
    
    pieces=kizeo_model_Pieces.objects.filter(kizeo_id=client_id)
    data = kizeo_model.objects.get(kizeo_id=client_id)
    return render(request, 'html/formK.html',{"data":data,"pieces":pieces,
                                              'Pieces_index_add':Pieces_index_add,
                                              'page_number':page_number,
                                              'client_id':client_id,
                                              'formKURL':formK111})




@login_required
def kizeo_form_Pieces(request,client_id,piece_id):
    user_ = request.user

    try:
        data=kizeo_model_Pieces.objects.get(kizeo_id=client_id,Pieces_index=piece_id)
    except:
        Pieces_index_add = kizeo_model_Pieces.objects.filter(kizeo_id=client_id).aggregate(Max('Pieces_index'))['Pieces_index__max']
        kizeo_model_Pieces.objects.create(kizeo_id=client_id,Pieces_index=piece_id)
    
    if "kize" in user_.role:
        if request.method == 'POST':
            
                        
            obj = kizeo_model_Pieces.objects.get(kizeo_id=client_id,Pieces_index=piece_id)  
                

            table_index=['Pieces_image_1',
                            'Pieces_image_2',
                            'Pieces_image_3'
                        ]
            for lis in table_index:
                img_get= request.FILES.get(lis)
                if img_get:
                    if hasattr(obj, lis):
                        if getattr(obj,lis):
                            if len(getattr(obj,lis)) > 0 :
                                os.remove(getattr(obj,lis).path)
                        # Check if the field exists in the model
                        setattr(obj, lis, img_get)
                        obj.save()  
            
            obj.Niveau = request.POST.get("Niveau")
            
            obj.Pieces = request.POST.get("Pieces")
            
            obj.Chauffage ="Non"
            Chauffage = request.POST.get("Chauffage")
            if Chauffage:
                obj.Chauffage ="Oui"
                
                
            Mansardee = request.POST.get("Mansardee")
            obj.Mansardee ="Non"
            
            obj.IF_mansardee = False
            if request.POST.get("IF_mansardee"):
                obj.IF_mansardee = True
                obj.HSF = request.POST.get("HSF")
                obj.HPP = request.POST.get("HPP")
                obj.LP = request.POST.get("LP")
                obj.S_rampants_1 = request.POST.get("S_rampants_1")
                obj.S_rampants_2 = request.POST.get("S_rampants_2")
                obj.save(update_fields=['HSF','HPP','LP','S_rampants_1','S_rampants_2'])
            

            



            
                
            
            obj.HSP = float(request.POST.get("HSP"))
            
            obj.Longueur = float(request.POST.get("Longueur"))
            obj.Largeur = float(request.POST.get("Largeur"))
            obj.Surface = float(request.POST.get("Longueur"))*float(request.POST.get("Largeur"))
            
            obj.Decrochement_Longueur = float(request.POST.get("Decrochement_Longueur"))
            obj.Decrochement_Largeur = float(request.POST.get("Decrochement_Largeur"))
            obj.Decrochement_Surface = float(request.POST.get("Decrochement_Longueur")) * float(request.POST.get("Decrochement_Largeur"))
            
            obj.Surface_nette = float(request.POST.get("Surface_nette"))
            
            obj.Menuiseries_F = request.POST.get("Menuiseries_F")
            obj.Menuiseries_L = float(request.POST.get("Menuiseries_L"))
            obj.Menuiseries_H = float(request.POST.get("Menuiseries_H"))
            obj.Menuiseries_N = int(request.POST.get("Menuiseries_N"))
            
            obj.save(update_fields=[
                                    'Niveau',
            
                                    'Pieces',
                                    'Chauffage',
                                    'Mansardee',
                                    'HSP',
                                    
                                    'Longueur',
                                    'Largeur',
                                    'Surface',
                                    
                                    'IF_mansardee',
                                    
                                    'Decrochement_Longueur',
                                    'Decrochement_Largeur',
                                    'Decrochement_Surface',
                                    
                                    'Surface_nette',
                                    
                                    'Menuiseries_F',
                                    'Menuiseries_L',
                                    'Menuiseries_H',
                                    'Menuiseries_N',
                ])
                                    
    data = kizeo_model_Pieces.objects.get(kizeo_id=client_id,Pieces_index=piece_id)
    
    return render(request, 'html/formK_Pieces.html',{'data':data})


@login_required
def kizeo_form_Pieces_delete(request,client_id,piece_id):
    user_= request.user
    user_role= user_.role
    if  "kize" in user_role:
        kizeo_model_Pieces.objects.get(kizeo_id=client_id,Pieces_index=piece_id).delete()
        return redirect(f"{formK}{client_id}/5")
    
    return redirect("/")
    
    
@login_required
def save_signature(request):
    if request.method == 'POST':
        return redirect("gggg")
        signature_data = request.POST.get('signature', '')

        # You can save the signature data to your model as an image field.
        # Here, assume you have a model named Signature with a signature field.
        obj = kizeo_model.objects.get(kizeo_id="4MLJMq2cj8")
        obj.objects.create(signature_image=signature_data)
        return JsonResponse({'message': 'Signature saved successfully.'})

    
@login_required   
def download_K_file(request,file_id):
    if kizeo_model.objects.filter(kizeo_id=file_id):
            pass
    else:
        kizeo_model.objects.create(kizeo_id=file_id)
    
    obj = kizeo_model.objects.get(kizeo_id=file_id)
    obj2 = kizeo_model_Pieces.objects.filter(kizeo_id=file_id)
    
    template_path = XLSX_TEMPLATE_FILE_PATH
    #template_path = 'ERapp\static\KiFile.xlsx'  # Provide the path to your template file
    workbook = openpyxl.load_workbook(template_path)
    
    worksheet1 = workbook["Données"]
    worksheet2 = workbook["Métré"]
    worksheet3 = workbook["Garde"]
    worksheet4 = workbook["Site"]
    
    ###################################### start WORK_SHEET_1
    ###################################### start WORK_SHEET_1
    ###################################### start WORK_SHEET_1
    i=2
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Nom_d_intervenant
    i+=1#3
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Date_de_visite
    i+=1#4
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Adresse
    i+=1#5
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Zip_Code
    i+=1#6
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_City
    i+=1#7
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Annee_de_construction
    i+=1#8
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Nom_client
    i+=1#9
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Tel_client
    i+=1#10
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Email
    i+=1#11
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = "obj."
    i+=1#12
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = "obj."
    i+=1#13
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Nombre_d_occupant
    i+=1#14
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Surface_TOTALE
    i+=1#15
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Preuve_Surface
    i+=1#16
    text_cell = worksheet1[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Surface_ajoute_depuis_moins_de_15_ans
    
    
    
    ###################################### end WORK_SHEET_1
    ###################################### end WORK_SHEET_1
    ###################################### end WORK_SHEET_1
    
    
    ###################################### start WORK_SHEET_2
    ###################################### start WORK_SHEET_2
    ###################################### start WORK_SHEET_2
    
    
    ### Pieces
    j=3
    for item in obj2:
        image_cell_1 = worksheet2[f'A{j}'] 
        image_cell_1.value = None
        o = item.Pieces_image_1
        if o:
            image_path_ = o.path  # Get the file path
            if default_storage.exists(image_path_):
                image_path1 = o.path
                img_1 = Image(image_path1)
                img_1.width =  185.28 #cell_width
                img_1.height = 140.16 # cell_height
                worksheet2.add_image(img_1, image_cell_1.coordinate)
            
        image_cell_1 = worksheet2[f'B{j}'] 
        image_cell_1.value = None
        o = item.Pieces_image_2
        if o:
            image_path_ = o.path  # Get the file path
            if default_storage.exists(image_path_):
                image_path1 = o.path
                img_1 = Image(image_path1)
                img_1.width =  185.28 #cell_width
                img_1.height = 140.16 # cell_height
                worksheet2.add_image(img_1, image_cell_1.coordinate)
                
        image_cell_1 = worksheet2[f'C{j}'] 
        image_cell_1.value = None
        o = item.Pieces_image_3
        if o:
            image_path_ = o.path  # Get the file path
            if default_storage.exists(image_path_):
                image_path1 = o.path
                img_1 = Image(image_path1)
                img_1.width =  185.28 #cell_width
                img_1.height = 140.16 # cell_height
                worksheet2.add_image(img_1, image_cell_1.coordinate)
                
                
        text_cell = worksheet2[f'D{j}'] 
        text_cell.value = item.Pieces
        text_cell = worksheet2[f'E{j}'] 
        text_cell.value = item.Chauffage
        
        
        text_cell = worksheet2[f'G{j}'] 
        text_cell.value = item.HSP
        
        text_cell = worksheet2[f'H{j}'] 
        text_cell.value = item.Longueur
        text_cell = worksheet2[f'I{j}'] 
        text_cell.value = item.Largeur
        text_cell = worksheet2[f'J{j}'] 
        text_cell.value = item.Surface
        
        text_cell = worksheet2[f'F{j}'] 
        text_cell.value = "Non"
        if item.IF_mansardee:
            text_cell.value = 'Oui'
            text_cell = worksheet2[f'K{j}'] 
            text_cell.value = item.HSF
            text_cell = worksheet2[f'L{j}'] 
            text_cell.value = item.HPP
            text_cell = worksheet2[f'M{j}'] 
            text_cell.value = item.LP
            text_cell = worksheet2[f'N{j}'] 
            text_cell.value = item.S_rampants_1
            text_cell = worksheet2[f'O{j}'] 
            text_cell.value = item.S_rampants_2
        
        text_cell = worksheet2[f'P{j}'] 
        text_cell.value = item.Decrochement_Longueur
        text_cell = worksheet2[f'Q{j}'] 
        text_cell.value = item.Decrochement_Largeur
        text_cell = worksheet2[f'R{j}'] 
        text_cell.value = item.Decrochement_Surface
        
        text_cell = worksheet2[f'S{j}'] 
        text_cell.value = item.Surface_nette
        
        text_cell = worksheet2[f'T{j}'] 
        text_cell.value = item.Menuiseries_F
        text_cell = worksheet2[f'U{j}'] 
        text_cell.value = item.Menuiseries_L
        text_cell = worksheet2[f'V{j}'] 
        text_cell.value = item.Menuiseries_H
        text_cell = worksheet2[f'W{j}'] 
        text_cell.value = item.Menuiseries_N
        
        j+=1
        
        
    ###################################### end WORK_SHEET_2
    ###################################### end WORK_SHEET_2
    ###################################### end WORK_SHEET_2
    
    
    ###################################### start WORK_SHEET_3
    ###################################### start WORK_SHEET_3
    ###################################### start WORK_SHEET_3
    image_cell_1 = worksheet3['B6'] 
    image_cell_1.value = None
    o = obj.Donnees_Generales_Preuve_Surface_Photo
    if o:
        image_path_ = o.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = o.path
            img_1 = Image(image_path1)
            img_1.width =  382.08 #cell_width 3.98
            img_1.height = 280.32 # cell_height 2.92
            worksheet3.add_image(img_1, image_cell_1.coordinate)
            
    i=26
    text_cell = worksheet3[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Nom_client
    i+=1
    text_cell = worksheet3[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Adresse
    i+=1
    text_cell = worksheet3[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_Zip_Code
    i+=1
    text_cell = worksheet3[f'B{i}']  
    text_cell.value = obj.Donnees_Generales_City
    
    text_cell = worksheet3['C33']  
    text_cell.value = obj.Donnees_Generales_Date_de_visite
    ###################################### end WORK_SHEET_3
    ###################################### end WORK_SHEET_3
    ###################################### end WORK_SHEET_3
    
    ################################### start WORK_SHEET_4
    ################################### start WORK_SHEET_4
    ################################### start WORK_SHEET_4
    text_cell = worksheet4['A7']
    text_cell.value =f"L'audit énergétique est réalisé une maison individuelle située à {obj.Donnees_Generales_Adresse}, {obj.Donnees_Generales_Zip_Code} {obj.Donnees_Generales_City}."
    
    #"https://google.com/maps?q={{data.latitude}},{{data.longitude}}"
    text_cell = worksheet4['A8']  
    #text_cell.value = "GPS"
    #text_cell.hyperlink  = f"https://google.com/maps?q={obj.latitude},{obj.longitude}"
    
    text_cell = worksheet4['E27']  
    text_cell.value = obj.Donnees_Generales_Annee_de_construction
    text_cell = worksheet4['E28']  
    text_cell.value = obj.Donnees_Generales_Etat_d_occupation
    text_cell = worksheet4['E29'] 
    text_cell.value = obj.Donnees_Generales_Surface_TOTALE
    
    
    text_cell = worksheet4['A33']  
    text_cell.value = obj.Facade_1_Orientation
    text_cell = worksheet4['A34']  
    text_cell.value = obj.Facade_1_Mitoyennete
    text_cell = worksheet4['A35'] 
    text_cell.value = f"Longueur = {obj.Facade_1_Longueur}"
    text_cell = worksheet4['A36']  
    text_cell.value = f"Hauteur = {obj.Facade_1_Hauteur}"
    text_cell = worksheet4['A37']  
    text_cell.value = f"Surface = {obj.Facade_1_Surface}"
    
    image_cell_1 = worksheet4['A38'] 
    image_cell_1.value = None
    #image_cell_1.hyperlink = None
    if obj.Facade_1_Photo_Principale:
        image_path_ = obj.Facade_1_Photo_Principale.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = obj.Facade_1_Photo_Principale.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    
    text_cell = worksheet4['E33']  
    text_cell.value = obj.Facade_2_Orientation
    text_cell = worksheet4['E34']  
    text_cell.value = obj.Facade_2_Mitoyennete
    text_cell = worksheet4['E35'] 
    text_cell.value = f"Longueur = {obj.Facade_2_Longueur}"
    text_cell = worksheet4['E36']  
    text_cell.value = f"Hauteur = {obj.Facade_2_Hauteur}"
    text_cell = worksheet4['E37']  
    text_cell.value = f"Surface = {obj.Facade_2_Surface}"
    
    image_cell_1 = worksheet4['E38']  
    image_cell_1.value = None
    if obj.Facade_2_Photo_Principale:
        image_path_ = obj.Facade_2_Photo_Principale.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = obj.Facade_2_Photo_Principale.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    
    text_cell = worksheet4['A46']  
    text_cell.value = obj.Facade_3_Orientation
    #text_cell = worksheet4['A4']  
    #text_cell.value = obj.Facade_3_Mitoyennete
    text_cell = worksheet4['A47'] 
    text_cell.value = f"Longueur = {obj.Facade_3_Longueur}"
    text_cell = worksheet4['A48']  
    text_cell.value = f"Hauteur = {obj.Facade_3_Hauteur}"
    text_cell = worksheet4['A49']  
    text_cell.value = f"Surface = {obj.Facade_3_Surface}"
    
    image_cell_1 = worksheet4['A50'] 
    image_cell_1.value = None
    #image_cell_1.hyperlink = None
    if obj.Facade_3_Photo_Principale:
        image_path_ = obj.Facade_3_Photo_Principale.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = obj.Facade_3_Photo_Principale.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    
    text_cell = worksheet4['E46']  
    text_cell.value = obj.Facade_4_Orientation
    #text_cell = worksheet4['E34']  
    #text_cell.value = obj.Facade_4_Mitoyennete
    text_cell = worksheet4['E47'] 
    text_cell.value = f"Longueur = {obj.Facade_4_Longueur}"
    text_cell = worksheet4['E48']  
    text_cell.value = f"Hauteur = {obj.Facade_4_Hauteur}"
    text_cell = worksheet4['E49']  
    text_cell.value = f"Surface = {obj.Facade_4_Surface}"
    
    image_cell_1 = worksheet4['E50'] 
    image_cell_1.value = None
    if obj.Facade_4_Photo_Principale:
        image_path_ = obj.Facade_4_Photo_Principale.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = obj.Facade_4_Photo_Principale.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    
    
    ### mur 1
    #text_cell1 = worksheet4.merge_cells('B60:E60') 
    #cell = text_cell1.cell(row=1, column=1) 
    text_cell = worksheet4['B60'] 
    text_cell.value = f"Murs - type 1 - {obj.Mur_1_Position}"
    text_cell = worksheet4['E61']  
    text_cell.value = obj.Mur_1_Composition
    text_cell = worksheet4['E62'] 
    text_cell.value = obj.Mur_1_Epaisseur_mur
    text_cell = worksheet4['E63']  
    text_cell.value = obj.Mur_1_Isolation
    text_cell = worksheet4['E64']  
    text_cell.value = obj.Mur_1_Epaisseur_isolant
    text_cell = worksheet4['E65']  
    text_cell.value = obj.Mur_1_Date_d_isolation
    text_cell = worksheet4['E66']  
    #text_cell.value = obj.
    
    image_cell_1 = worksheet4['B67'] 
    image_cell_1.value = None
    if obj.Mur_1_Photo_mur:
        image_path_ = obj.Mur_1_Photo_mur.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = obj.Mur_1_Photo_mur.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    ### mur 2
    #text_cell1 = worksheet4.merge_cells('B60:E60') 
    #cell = text_cell1.cell(row=1, column=1) 
    text_cell = worksheet4['B76'] 
    text_cell.value = f"Murs - type 2 - {obj.Mur_2_Position}"
    text_cell = worksheet4['E77']  
    text_cell.value = obj.Mur_2_Composition
    text_cell = worksheet4['E78'] 
    text_cell.value = obj.Mur_2_Epaisseur_mur
    text_cell = worksheet4['E79']  
    text_cell.value = obj.Mur_2_Isolation
    text_cell = worksheet4['E80']  
    text_cell.value = obj.Mur_2_Epaisseur_isolant
    text_cell = worksheet4['E81']  
    text_cell.value = obj.Mur_2_Date_d_isolation
    text_cell = worksheet4['E82']  
    #text_cell.value = obj.
    
    image_cell_1 = worksheet4['B83'] 
    image_cell_1.value = None
    if obj.Mur_2_Photo_mur:
        image_path_ = obj.Mur_2_Photo_mur.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = obj.Mur_2_Photo_mur.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    
    
    ### Plancher bas type1
    text_cell = worksheet4['E92'] 
    text_cell.value = obj.Plancher_bas_1_Position
    text_cell = worksheet4['E93']  
    text_cell.value = obj.Plancher_bas_1_Composition
    text_cell = worksheet4['E94'] 
    text_cell.value = obj.Plancher_bas_1_Isolation
    text_cell = worksheet4['E95']  
    text_cell.value = obj.Plancher_bas_1_Epaisseur_isolant
    text_cell = worksheet4['E96']  
    text_cell.value = obj.Plancher_bas_1_Date_d_isolation
    text_cell = worksheet4['E97']  
    text_cell.value = obj.Plancher_bas_1_Surface
    text_cell = worksheet4['E98']  
    #text_cell.value = obj.
    
    image_cell_1 = worksheet4['B99'] 
    image_cell_1.value = None
    if obj.Plancher_bas_1_Photo_plancher_bas:
        image_path_ = obj.Plancher_bas_1_Photo_plancher_bas.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = obj.Plancher_bas_1_Photo_plancher_bas.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    ################################### end Facade
    
    ### Plancher bas type1
    j=108
    for i in [obj.Plancher_bas_2_Position,obj.Plancher_bas_2_Composition,obj.Plancher_bas_2_Isolation
              ,obj.Plancher_bas_2_Epaisseur_isolant,obj.Plancher_bas_2_Date_d_isolation,obj.Plancher_bas_2_Surface]:
        
        text_cell = worksheet4[f'E{j}'] 
        text_cell.value = i
        j+=1
        
    image_cell_1 = worksheet4['B115'] 
    image_cell_1.value = None
    if obj.Plancher_bas_2_Photo_plancher_bas:
        image_path_ = obj.Plancher_bas_2_Photo_plancher_bas.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = obj.Plancher_bas_2_Photo_plancher_bas.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    ### Plancher haut type1
    j=124
    for i in [obj.Plancher_Haut_1_Type,obj.Plancher_Haut_1_Composition,obj.Plancher_Haut_1_Isolation
              ,obj.Plancher_Haut_1_Epaisseur_isolant,obj.Plancher_Haut_1_Date_d_isolation,obj.Plancher_Haut_1_Surface]:
        
        text_cell = worksheet4[f'E{j}'] 
        text_cell.value = i
        j+=1
        
    objimage =obj.Plancher_Haut_1_Photo_plancher_bas
    image_cell_1 = worksheet4['B131'] 
    image_cell_1.value = None
    if objimage:
        image_path_ =objimage.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = objimage.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    ### Plancher haut type 2
    j=140
    for i in [obj.Plancher_Haut_2_Type,obj.Plancher_Haut_2_Composition,obj.Plancher_Haut_2_Isolation
              ,obj.Plancher_Haut_2_Epaisseur_isolant,obj.Plancher_Haut_2_Date_d_isolation,obj.Plancher_Haut_2_Surface]:
        
        text_cell = worksheet4[f'E{j}'] 
        text_cell.value = i
        j+=1
        
    objimage =obj.Plancher_Haut_2_Photo_plancher_bas
    image_cell_1 = worksheet4['B147'] 
    image_cell_1.value = None
    if objimage:
        image_path_ =objimage.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = objimage.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    ### Fenêtre 1
    text_cell = worksheet4[f'B155'] 
    text_cell.value = obj.Fenetre_type_1_Menuiserie
    j=156
    
    text_cell = worksheet4[f'E{j}'] 
    text_cell.value = obj.Fenetre_type_1_Materiaux
    j+1
    text_cell = worksheet4[f'E{j}'] 
    text_cell.value = obj.Fenetre_type_2_Type_de_vitrage
    j+1
    text_cell = worksheet4[f'E{j}'] 
    text_cell.value = obj.Fenetre_type_2_Volets
    j+1
    text_cell = worksheet4[f'E{j}'] 
    text_cell.value = obj.Fenetre_type_2_Nombre
    for i in [obj.Fenetre_type_1_Materiaux,obj.Fenetre_type_1_Type_de_vitrage,
              obj.Fenetre_type_1_Volets,obj.Fenetre_type_1_Nombre]:
        
        #text_cell = worksheet4[f'E{j}'] 
        #text_cell.value = str(i)
        j+=1
        
    objimage =obj.Fenetre_type_1_Photo
    image_cell_1 = worksheet4['B160'] 
    image_cell_1.value = None
    if objimage:
        image_path_ =objimage.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = objimage.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    ### Fenêtre 2
    j=169
    
    text_cell = worksheet4[f'B{j}'] 
    text_cell.value = f" {obj.Fenetre_type_1_Menuiserie}"
    j+=1
    for i in [obj.Fenetre_type_1_Materiaux,obj.Fenetre_type_2_Type_de_vitrage,
              obj.Fenetre_type_2_Volets,obj.Fenetre_type_2_Nombre]:
        
        #text_cell = worksheet4[f'E{j}'] 
        #text_cell.value = i
        j+=1
    
    
    
    
    objimage =obj.Fenetre_type_2_Photo
    image_cell_1 = worksheet4['B174'] 
    image_cell_1.value = None
    if objimage:
        image_path_ =objimage.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = objimage.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    ### Porte 1
    j=183
    text_cell = worksheet4[f'E{j}'] 
    text_cell.value = obj.Porte_1_Materiaux
    j+=1
    text_cell = worksheet4[f'E{j}'] 
    text_cell.value = obj.Porte_1_Type_porte
    j+=1
    text_cell = worksheet4[f'E{j}'] 
    text_cell.value = obj.Porte_1_Nombre
    j+=1    
    objimage =obj.Porte_1_Photo_porte
    image_cell_1 = worksheet4['B186'] 
    image_cell_1.value = None
    if objimage:
        image_path_ =objimage.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = objimage.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
            
            
    ### Porte 2
    j=195
    text_cell = worksheet4[f'E{j}'] 
    text_cell.value = obj.Porte_2_Materiaux
    j+=1
    text_cell = worksheet4[f'E{j}'] 
    text_cell.value = obj.Porte_2_Type_porte
    j+=1
    text_cell = worksheet4[f'E{j}'] 
    text_cell.value = obj.Porte_2_Nombre
    j+=1
    
    objimage =obj.Porte_2_Photo_porte
    image_cell_1 = worksheet4[f'B{j}'] 
    image_cell_1.value = None
    if objimage:
        image_path_ =objimage.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = objimage.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    ### Cauffage 
    j=209
    text_cell = worksheet4[f'D{j}'] 
    text_cell.value = obj.Cauffage_systeme
    j+=1
    text_cell = worksheet4[f'D{j}'] 
    text_cell.value = 'obj.'
    j+=1
    text_cell = worksheet4[f'D{j}'] 
    text_cell.value = obj.Cauffage_annee_de_mise_en_oeuvre
    j+=1
    text_cell = worksheet4[f'D{j}'] 
    text_cell.value = "obj."
    j+=1
    text_cell = worksheet4[f'D{j}'] 
    text_cell.value = "obj."
    j+=1
    text_cell = worksheet4[f'D{j}'] 
    text_cell.value = obj.Cauffage_type_de_regulation
    j+=1
    
    #objimage =obj.
    #image_cell_1 = worksheet4[f'B{j}'] 
    #image_cell_1.value = None
    #if objimage:
    #    image_path_ =objimage.path  # Get the file path
    #    if default_storage.exists(image_path_):
    #        image_path1 = objimage.path
    #        img_1 = Image(image_path1)
    #        img_1.width =  185.28 #cell_width
    #        img_1.height = 140.16 # cell_height
    #        worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    ###  ECS
    j=236
    text_cell = worksheet4[f'D{j}'] 
    text_cell.value = obj.ECS_type
    j+=1
    text_cell = worksheet4[f'D{j}'] 
    text_cell.value = "obj."
    j+=1
    objimage =obj.ECS_photo_appoint
    image_cell_1 = worksheet4[f'B{j}'] 
    image_cell_1.value = None
    if objimage:
        image_path_ =objimage.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = objimage.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    text_cell = worksheet4['D245'] 
    text_cell.value = obj.ECS_system_d_appoint
    
    
    ###  Ventilation
    j=249
    text_cell = worksheet4[f'D{j}'] 
    text_cell.value = obj.Ventilation_type
    j+=1
    
    objimage =obj.Ventilation_photo_ventilation
    image_cell_1 = worksheet4[f'B{j}'] 
    image_cell_1.value = None
    if objimage:
        image_path_ =objimage.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = objimage.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    ###  Refroidissement
    j=249
    text_cell = worksheet4[f'D{j}'] 
    text_cell.value = obj.Refroidissement_type
    j+=1
    
    #objimage =obj.Ventilation_photo_ventilation
    #image_cell_1 = worksheet4[f'B{j}'] 
    #image_cell_1.value = None
    #if objimage:
    #    image_path_ =objimage.path  # Get the file path
    #    if default_storage.exists(image_path_):
    #        image_path1 = objimage.path
    #        img_1 = Image(image_path1)
    #        img_1.width =  185.28 #cell_width
    #        img_1.height = 140.16 # cell_height
    #        worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    
    ###  Compteur Electrique
    j=269
    text_cell = worksheet4[f'D{j}'] 
    text_cell.value = obj.Compteur_Electrique_Puissance_souscrite
    j+=1
    text_cell = worksheet4[f'D{j}'] 
    text_cell.value = obj.Compteur_Electrique_type
    j+=1
    
    objimage =obj.Compteur_Electrique_photo_compteur
    image_cell_1 = worksheet4[f'B{j}'] 
    image_cell_1.value = None
    if objimage:
        image_path_ =objimage.path  # Get the file path
        if default_storage.exists(image_path_):
            image_path1 = objimage.path
            img_1 = Image(image_path1)
            img_1.width =  185.28 #cell_width
            img_1.height = 140.16 # cell_height
            worksheet4.add_image(img_1, image_cell_1.coordinate)
    
    ### Pieces
    j=284
    for item in obj2:
        
        text_cell = worksheet4[f'A{j}'] 
        text_cell.value = item.Pieces
        text_cell = worksheet4[f'B{j}'] 
        text_cell.value = item.Chauffage
        text_cell = worksheet4[f'C{j}'] 
        text_cell.value = item.Niveau
        text_cell = worksheet4[f'D{j}'] 
        text_cell.value = item.HSP
        text_cell = worksheet4[f'F{j}'] 
        text_cell.value = F"{item.Longueur}x{item.Largeur}"
        text_cell = worksheet4[f'G{j}'] 
        text_cell.value = item.Surface
        j+=1
    ######################################   END WORK_SHEET_4
    ######################################   END WORK_SHEET_4
    ######################################   END WORK_SHEET_4
    
        ###################################################### save process #######################################################
    # Save the updated XLSX file temporarily
    # Generate the XLSX content as bytes
# Create a BytesIO object to store the XLSX content
    xlsx_content = BytesIO()

    # Save the workbook to the BytesIO object
    workbook.save(xlsx_content)

    # Seek to the beginning of the BytesIO object
    xlsx_content.seek(0)

    # Create an HttpResponse with the XLSX content for download
    response = HttpResponse(xlsx_content.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Kizeo.xlsx"'

    return response









def generate_pdf(template_path, context):
    # Load the template
    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template(template_path)

    # Render the template with the provided context
    html_content = template.render(**context)

    # Generate PDF from HTML content
    pdf_buffer = BytesIO()
    p = canvas.Canvas(pdf_buffer, pagesize=letter)

    # Draw HTML content on the PDF
    p.drawString(100, 700, html_content)  # Adjust the position as needed

    # Reset the buffer pointer to the beginning
    pdf_buffer.seek(0)

    return pdf_buffer

if __name__ == "__main__ _ _":
    # Define the template path and context
    template_path = "template.html"
    context = {
        "title": "My PDF",
        "content": "Hello, this is a sample PDF!",
        "my_integer": 42,
        "my_string": "This is a string",
        "my_float": 3.14,
        "image_path": "path/to/your/image.jpg"  # Replace with the actual path
    }

    # Generate the PDF
    pdf_buffer = generate_pdf(template_path, context)

    # Save the PDF to a file
    with open("output.pdf", "wb") as output_file:
        output_file.write(pdf_buffer.read())